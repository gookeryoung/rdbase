"""数据查询与 CRUD 模块.

提供表数据浏览的查询与计数能力，以及行的增删改查（CRUD）能力。
基于 SQLAlchemy ``text()`` 手动构造 SQL，通过列名白名单校验防 SQL 注入。

设计要点：
- ``query_table_rows``: 查询表数据，返回 ``(rows, total)``
- ``count_table_rows``: 统计满足筛选条件的行数
- ``insert_row``/``update_row``/``delete_row``/``get_row``: 单行 CRUD，主键反查定位
- ``execute_sql``: 执行任意 SQL，自动区分 SELECT（返回结果集）与 DDL/DML（返回影响行数）
- ``explain_sql``: 多方言执行计划（SQLite ``EXPLAIN QUERY PLAN``、MySQL/PG ``EXPLAIN`` 可选 ``ANALYZE``）
- ``iter_table_rows``/``rows_to_csv``/``rows_to_sql``/``export_excel``: 流式导出（CSV/SQL/Excel）
- ``parse_csv_upload``/``parse_excel_upload``/``import_rows``: 流式导入（CSV/Excel，事务批量插入）
- 标识符引用：MySQL 反引号、PG/SQLite 双引号（独立实现，避免跨模块依赖）
- 白名单：通过 SQLAlchemy inspect 获取列名集合，校验 ``columns``/``order_by``/``filters``/``values``/``pk``
- SQLite schema 强制 None
- 乐观锁极简方案：UPDATE/DELETE 影响 0 行抛 ``QueryError``，不做 version 列推断
- 所有写操作在单个 ``engine.begin()`` 事务内
"""

from __future__ import annotations

import csv
import io
import itertools
import json
import time
from collections.abc import Iterator
from datetime import date, datetime
from datetime import time as dtime
from typing import Any, cast

from sqlalchemy import inspect, text
from sqlalchemy.engine import Engine
from sqlalchemy.exc import SQLAlchemyError

from apps.datasources.models import EngineType


class QueryError(ValueError):
    """查询错误（如列名非法、操作符不支持、表不存在等）."""


# 支持的筛选操作符 → SQL 比较运算符
# ``in`` 操作符需要展开多个占位符，单独处理
_COMPARATORS: frozenset[str] = frozenset({"eq", "ne", "gt", "lt", "ge", "le", "like", "in"})


def _quote_ident(name: str, dialect: str) -> str:
    """标识符引用（MySQL 用反引号，其他用双引号）."""
    if dialect == EngineType.MYSQL:
        return f"`{name}`"
    return f'"{name}"'


def _format_table_ref(table_name: str, schema: str | None, dialect: str) -> str:
    """生成表引用（含 schema 前缀）.

    SQLite 不支持 schema 限定，强制忽略 schema_name。
    """
    if schema and dialect != EngineType.SQLITE:
        return f"{_quote_ident(schema, dialect)}.{_quote_ident(table_name, dialect)}"
    return _quote_ident(table_name, dialect)


def _resolve_schema(engine: Engine, schema: str | None) -> str | None:
    """SQLite 不支持 schema 概念，强制返回 None；其他方言原样返回."""
    if engine.dialect.name == EngineType.SQLITE:
        return None
    return schema


def get_column_names(engine: Engine, table_name: str, schema: str | None) -> list[str]:
    """通过 SQLAlchemy inspect 获取表的列名列表.

    Raises:
        QueryError: 表不存在或反射失败。
    """
    insp = inspect(engine)
    effective_schema = _resolve_schema(engine, schema)
    try:
        columns = insp.get_columns(table_name, schema=effective_schema)
    except SQLAlchemyError as exc:
        raise QueryError(f"无法读取表 {table_name} 的列信息: {exc}") from None
    return [cast(str, c["name"]) for c in columns]


def _validate_columns(allowed: set[str], columns: list[str] | None) -> list[str]:
    """校验列名是否在白名单内."""
    if not columns:
        return []
    for col in columns:
        if col not in allowed:
            raise QueryError(f"非法列名: {col}")
    return columns


def _validate_order_by(allowed: set[str], order_by: str | None) -> str | None:
    """校验排序字段."""
    if not order_by:
        return None
    if order_by not in allowed:
        raise QueryError(f"非法排序字段: {order_by}")
    return order_by


def _validate_filter_columns(allowed: set[str], filters: dict[str, dict[str, Any]]) -> None:
    """校验筛选条件中的列名与操作符."""
    for col, cond in filters.items():
        if col not in allowed:
            raise QueryError(f"非法筛选列名: {col}")
        op = cond.get("op")
        if op not in _COMPARATORS:
            raise QueryError(f"不支持的操作符: {op}")


def _build_where_clause(
    filters: dict[str, dict[str, Any]],
    dialect: str,
) -> tuple[str, dict[str, Any]]:
    """构造 WHERE 子句与参数.

    Args:
        filters: 筛选条件，格式 ``{列名: {"op": "eq/ne/gt/lt/ge/le/like/in", "val": ...}}``.
        dialect: 方言名。

    Returns:
        ``(where_sql, params)``：``where_sql`` 不含 ``WHERE`` 关键字，空字符串表示无条件。
    """
    clauses: list[str] = []
    params: dict[str, Any] = {}
    for i, (col, cond) in enumerate(filters.items()):
        op = cond["op"]
        val = cond.get("val")
        col_ref = _quote_ident(col, dialect)
        param_base = f"f{i}"
        if op == "in":
            if not isinstance(val, list) or not val:
                raise QueryError("in 操作符的值须为非空列表")
            placeholders = ", ".join(f":{param_base}_{j}" for j in range(len(val)))
            clauses.append(f"{col_ref} IN ({placeholders})")
            for j, v in enumerate(val):
                params[f"{param_base}_{j}"] = v
        else:
            op_map = {
                "eq": "=",
                "ne": "<>",
                "gt": ">",
                "lt": "<",
                "ge": ">=",
                "le": "<=",
                "like": "LIKE",
            }
            clauses.append(f"{col_ref} {op_map[op]} :{param_base}")
            params[param_base] = val
    return " AND ".join(clauses), params


def query_table_rows(  # noqa: PLR0913, PLR0917
    engine: Engine,
    table_name: str,
    schema: str | None = None,
    columns: list[str] | None = None,
    page: int = 1,
    page_size: int = 20,
    order_by: str | None = None,
    order_dir: str = "asc",
    filters: dict[str, dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], int]:
    """查询表数据，返回 ``(行列表, 总数)``.

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 表名。
        schema: Schema 名（SQLite 强制 None）。
        columns: 显式指定的列名列表；None 或空表示所有列。
        page: 页码，从 1 开始。
        page_size: 每页行数。
        order_by: 排序字段；None 表示不排序。
        order_dir: 排序方向，``asc`` 或 ``desc``。
        filters: 筛选条件，格式 ``{列名: {"op": "eq/ne/gt/lt/ge/le/like/in", "val": ...}}``。

    Returns:
        ``(rows, total)``：``rows`` 为 dict 列表（键为列名），``total`` 为满足筛选条件的总行数。

    Raises:
        QueryError: 列名非法、操作符不支持、表不存在、分页参数非法等。
    """
    if page < 1:
        raise QueryError("page 须 >= 1")
    if page_size < 1:
        raise QueryError("page_size 须 >= 1")
    if order_dir not in ("asc", "desc"):
        raise QueryError("order_dir 须为 asc 或 desc")

    filters = filters or {}
    dialect = engine.dialect.name
    effective_schema = _resolve_schema(engine, schema)
    table_ref = _format_table_ref(table_name, effective_schema, dialect)

    allowed = set(get_column_names(engine, table_name, schema))
    selected_columns = _validate_columns(allowed, columns)
    _validate_order_by(allowed, order_by)
    _validate_filter_columns(allowed, filters)

    # SELECT 列：显式列出避免暴露非预期列；为空时用 *
    if selected_columns:
        select_cols = ", ".join(_quote_ident(c, dialect) for c in selected_columns)
    else:
        select_cols = "*"

    where_sql, params = _build_where_clause(filters, dialect)
    where_clause = f" WHERE {where_sql}" if where_sql else ""

    # 排序
    order_clause = ""
    if order_by:
        order_clause = f" ORDER BY {_quote_ident(order_by, dialect)} {order_dir.upper()}"

    # 分页（SQLite/MySQL/PG 均支持 LIMIT/OFFSET）
    offset = (page - 1) * page_size
    limit_clause = " LIMIT :_page_size OFFSET :_offset"
    params["_page_size"] = page_size
    params["_offset"] = offset

    select_sql = f"SELECT {select_cols} FROM {table_ref}{where_clause}{order_clause}{limit_clause}"
    count_sql = f"SELECT COUNT(*) FROM {table_ref}{where_clause}"

    with engine.connect() as conn:
        rows_result = conn.execute(text(select_sql), params)
        rows = [cast("dict[str, Any]", dict(row._mapping)) for row in rows_result.fetchall()]
        # 在同一连接内统计总数（避免重复反射）
        total = cast(int, conn.execute(text(count_sql), params).scalar())

    return rows, total


def count_table_rows(
    engine: Engine,
    table_name: str,
    schema: str | None = None,
    filters: dict[str, dict[str, Any]] | None = None,
) -> int:
    """统计满足筛选条件的行数.

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 表名。
        schema: Schema 名（SQLite 强制 None）。
        filters: 筛选条件，同 :func:`query_table_rows`。

    Returns:
        总行数。

    Raises:
        QueryError: 列名非法、操作符不支持、表不存在等。
    """
    filters = filters or {}
    dialect = engine.dialect.name
    effective_schema = _resolve_schema(engine, schema)
    table_ref = _format_table_ref(table_name, effective_schema, dialect)

    allowed = set(get_column_names(engine, table_name, schema))
    _validate_filter_columns(allowed, filters)

    where_sql, params = _build_where_clause(filters, dialect)
    where_clause = f" WHERE {where_sql}" if where_sql else ""
    count_sql = f"SELECT COUNT(*) FROM {table_ref}{where_clause}"

    with engine.connect() as conn:
        return cast(int, conn.execute(text(count_sql), params).scalar())


# ============================================================
# P4-2 行 CRUD
# ============================================================


def get_pk_columns(engine: Engine, table_name: str, schema: str | None) -> list[str]:
    """通过 SQLAlchemy inspect 获取表的主键列名列表.

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 表名。
        schema: Schema 名（SQLite 强制 None）。

    Returns:
        主键列名列表（无主键时返回空列表）。

    Raises:
        QueryError: 表不存在或反射失败。
    """
    insp = inspect(engine)
    effective_schema = _resolve_schema(engine, schema)
    try:
        pk = insp.get_pk_constraint(table_name, schema=effective_schema)
    except SQLAlchemyError as exc:
        raise QueryError(f"无法读取表 {table_name} 的主键信息: {exc}") from None
    return [cast(str, c) for c in pk.get("constrained_columns", [])]


def _build_pk_where_clause(
    pk: dict[str, Any],
    dialect: str,
    param_prefix: str = "p",
) -> tuple[str, dict[str, Any]]:
    """构造主键 WHERE 子句与参数（不含 ``WHERE`` 关键字）.

    Args:
        pk: 主键列名 → 值的 dict。
        dialect: 方言名。
        param_prefix: 参数名前缀，避免与 SET 子句参数冲突。

    Returns:
        ``(where_sql, params)``：空主键时 ``where_sql`` 为空字符串。
    """
    clauses: list[str] = []
    params: dict[str, Any] = {}
    for i, (col, val) in enumerate(pk.items()):
        clauses.append(f"{_quote_ident(col, dialect)} = :{param_prefix}{i}")
        params[f"{param_prefix}{i}"] = val
    return " AND ".join(clauses), params


def _select_row_by_pk(
    conn: Any,
    table_ref: str,
    dialect: str,
    pk: dict[str, Any],
) -> dict[str, Any] | None:
    """在同一连接内按主键查单行.

    Args:
        conn: SQLAlchemy 连接（已在事务内）。
        table_ref: 已格式化的表引用。
        dialect: 方言名。
        pk: 主键列名 → 值的 dict。

    Returns:
        行数据 dict；不存在时返回 None。
    """
    where_sql, params = _build_pk_where_clause(pk, dialect)
    select_sql = f"SELECT * FROM {table_ref} WHERE {where_sql} LIMIT 1"
    row = conn.execute(text(select_sql), params).fetchone()
    if row is None:
        return None
    return cast("dict[str, Any]", dict(row._mapping))


def insert_row(
    engine: Engine,
    table_name: str,
    schema: str | None,
    values: dict[str, Any],
) -> dict[str, Any]:
    """插入单行并返回插入后的行（含自增主键回填）.

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 表名。
        schema: Schema 名（SQLite 强制 None）。
        values: 列名 → 值的 dict。

    Returns:
        插入后的完整行数据（含自增主键回填）。

    Raises:
        QueryError: 列名非法、表不存在、主键缺失且无法回填等。
        SQLAlchemyError: 底层 SQL 执行失败（如约束冲突）。
    """
    if not values:
        raise QueryError("values 不能为空")

    allowed = set(get_column_names(engine, table_name, schema))
    for col in values:
        if col not in allowed:
            raise QueryError(f"非法列名: {col}")

    dialect = engine.dialect.name
    effective_schema = _resolve_schema(engine, schema)
    table_ref = _format_table_ref(table_name, effective_schema, dialect)

    cols = list(values.keys())
    col_refs = ", ".join(_quote_ident(c, dialect) for c in cols)
    placeholders = ", ".join(f":v{i}" for i in range(len(cols)))
    params = {f"v{i}": val for i, val in enumerate(values.values())}
    insert_sql = f"INSERT INTO {table_ref} ({col_refs}) VALUES ({placeholders})"

    pk_cols = get_pk_columns(engine, table_name, schema)
    if not pk_cols:
        # 无主键表：直接返回传入的 values（无法回查定位）
        with engine.begin() as conn:
            conn.execute(text(insert_sql), params)
        return dict(values)

    # 构造主键值：优先用 values 中提供的，缺失的用 lastrowid（单列自增主键场景）
    pk_values: dict[str, Any] = {}
    for pk_col in pk_cols:
        if pk_col in values:
            pk_values[pk_col] = values[pk_col]
    missing_pk = [c for c in pk_cols if c not in pk_values]
    # 多列主键场景：只要任一主键列缺失就抛错（lastrowid 仅能回填单列，无法定位多列主键）
    if missing_pk and len(pk_cols) > 1:
        raise QueryError(
            f"多列主键 {pk_cols} 须显式提供全部主键列，缺失: {missing_pk}",
        )

    with engine.begin() as conn:
        result = conn.execute(text(insert_sql), params)
        if missing_pk:
            # 单列自增主键场景：用 lastrowid 回填
            lastrowid = getattr(result, "lastrowid", None)
            if lastrowid is None:
                raise QueryError(
                    f"无法获取自增主键 {missing_pk}（dialect={dialect} 不支持 lastrowid）",
                )
            pk_values[missing_pk[0]] = lastrowid
        row = _select_row_by_pk(conn, table_ref, dialect, pk_values)
        if row is None:
            raise QueryError("插入后回查失败：行不存在")
        return row


def update_row(  # noqa: PLR0912
    engine: Engine,
    table_name: str,
    schema: str | None,
    pk: dict[str, Any],
    values: dict[str, Any],
) -> dict[str, Any]:
    """按主键更新单行并返回更新后的行.

    乐观锁极简方案：UPDATE 影响 0 行抛 ``QueryError``（行不存在或已被修改）。

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 表名。
        schema: Schema 名（SQLite 强制 None）。
        pk: 主键列名 → 值的 dict（多列主键支持）。
        values: 待更新列名 → 值的 dict。

    Returns:
        更新后的完整行数据。

    Raises:
        QueryError: 主键为空、values 为空、列名非法、主键列出现在 values 中、行不存在等。
        SQLAlchemyError: 底层 SQL 执行失败。
    """
    if not pk:
        raise QueryError("主键不能为空")
    if not values:
        raise QueryError("更新值不能为空")

    allowed = set(get_column_names(engine, table_name, schema))
    for col in pk:
        if col not in allowed:
            raise QueryError(f"非法主键列名: {col}")
    for col in values:
        if col not in allowed:
            raise QueryError(f"非法列名: {col}")

    pk_cols = get_pk_columns(engine, table_name, schema)
    if not pk_cols:
        raise QueryError("该表无主键，无法定位行")
    if set(pk.keys()) != set(pk_cols):
        raise QueryError(
            f"主键列不匹配：期望 {pk_cols}，实际 {list(pk.keys())}",
        )
    # 主键列不能在 values 中（避免修改主键）
    for col in values:
        if col in pk:
            raise QueryError(f"不能修改主键列: {col}")

    dialect = engine.dialect.name
    effective_schema = _resolve_schema(engine, schema)
    table_ref = _format_table_ref(table_name, effective_schema, dialect)

    set_clauses: list[str] = []
    params: dict[str, Any] = {}
    for i, (col, val) in enumerate(values.items()):
        set_clauses.append(f"{_quote_ident(col, dialect)} = :s{i}")
        params[f"s{i}"] = val
    where_sql, where_params = _build_pk_where_clause(pk, dialect)
    params.update(where_params)
    update_sql = f"UPDATE {table_ref} SET {', '.join(set_clauses)} WHERE {where_sql}"

    with engine.begin() as conn:
        result = conn.execute(text(update_sql), params)
        if result.rowcount == 0:
            raise QueryError("行不存在或已被修改")
        row = _select_row_by_pk(conn, table_ref, dialect, pk)
        if row is None:
            raise QueryError("更新后回查失败：行不存在")
        return row


def delete_row(
    engine: Engine,
    table_name: str,
    schema: str | None,
    pk: dict[str, Any],
) -> None:
    """按主键删除单行.

    乐观锁极简方案：DELETE 影响 0 行抛 ``QueryError``（行不存在或已被删除）。

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 表名。
        schema: Schema 名（SQLite 强制 None）。
        pk: 主键列名 → 值的 dict（多列主键支持）。

    Raises:
        QueryError: 主键为空、列名非法、表无主键、主键列不匹配、行不存在等。
        SQLAlchemyError: 底层 SQL 执行失败。
    """
    if not pk:
        raise QueryError("主键不能为空")

    allowed = set(get_column_names(engine, table_name, schema))
    for col in pk:
        if col not in allowed:
            raise QueryError(f"非法主键列名: {col}")

    pk_cols = get_pk_columns(engine, table_name, schema)
    if not pk_cols:
        raise QueryError("该表无主键，无法定位行")
    if set(pk.keys()) != set(pk_cols):
        raise QueryError(
            f"主键列不匹配：期望 {pk_cols}，实际 {list(pk.keys())}",
        )

    dialect = engine.dialect.name
    effective_schema = _resolve_schema(engine, schema)
    table_ref = _format_table_ref(table_name, effective_schema, dialect)
    where_sql, params = _build_pk_where_clause(pk, dialect)
    delete_sql = f"DELETE FROM {table_ref} WHERE {where_sql}"

    with engine.begin() as conn:
        result = conn.execute(text(delete_sql), params)
        if result.rowcount == 0:
            raise QueryError("行不存在或已被删除")


def get_row(
    engine: Engine,
    table_name: str,
    schema: str | None,
    pk: dict[str, Any],
) -> dict[str, Any] | None:
    """按主键查单行.

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 表名。
        schema: Schema 名（SQLite 强制 None）。
        pk: 主键列名 → 值的 dict（多列主键支持）。

    Returns:
        行数据 dict；不存在时返回 None。

    Raises:
        QueryError: 主键为空、列名非法、表无主键、主键列不匹配等。
    """
    if not pk:
        raise QueryError("主键不能为空")

    allowed = set(get_column_names(engine, table_name, schema))
    for col in pk:
        if col not in allowed:
            raise QueryError(f"非法主键列名: {col}")

    pk_cols = get_pk_columns(engine, table_name, schema)
    if not pk_cols:
        raise QueryError("该表无主键，无法定位行")
    if set(pk.keys()) != set(pk_cols):
        raise QueryError(
            f"主键列不匹配：期望 {pk_cols}，实际 {list(pk.keys())}",
        )

    dialect = engine.dialect.name
    effective_schema = _resolve_schema(engine, schema)
    table_ref = _format_table_ref(table_name, effective_schema, dialect)
    with engine.connect() as conn:
        return _select_row_by_pk(conn, table_ref, dialect, pk)


# ============================================================
# P4-3 SQL 查询控制台
# ============================================================


# 只读语句前缀（不区分大小写、去前导空白后判断）：这些语句不修改数据，
# viewer 角色可执行；其他语句（INSERT/UPDATE/DELETE/DDL）须 designer+。
_READ_ONLY_PREFIXES: tuple[str, ...] = ("select", "with", "show", "describe", "desc", "explain")

# 执行计划支持的方言 → EXPLAIN 语法模板
# 模板占位 ``{sql}`` 为原始 SQL，``{analyze}`` 在 PG/MySQL 8.0+ 用于追加 ANALYZE 关键字。
_EXPLAIN_TEMPLATES: dict[str, str] = {
    EngineType.SQLITE: "EXPLAIN QUERY PLAN {sql}",
    EngineType.POSTGRESQL: "EXPLAIN {analyze}{sql}",
    EngineType.MYSQL: "EXPLAIN {analyze}{sql}",
}


def _strip_sql(sql: str) -> str:
    """去除 SQL 末尾的分号与首尾空白，便于前缀判断与 EXPLAIN 拼接.

    Args:
        sql: 原始 SQL 字符串。

    Returns:
        去除首尾空白与末尾分号后的 SQL。

    Raises:
        QueryError: SQL 为空或仅含空白/分号。
    """
    cleaned = sql.strip().rstrip(";").strip()
    if not cleaned:
        raise QueryError("SQL 不能为空")
    return cleaned


def _is_read_only(sql: str) -> bool:
    """判断 SQL 是否为只读语句（SELECT/WITH/SHOW/DESCRIBE/DESC/EXPLAIN）.

    Args:
        sql: 已去空白与末尾分号的 SQL。

    Returns:
        True 表示只读；False 表示 DDL/DML（写操作）。
    """
    lowered = sql.lower()
    return lowered.startswith(_READ_ONLY_PREFIXES)


def execute_sql(
    engine: Engine,
    sql: str,
    *,
    read_only: bool = False,
) -> dict[str, Any]:
    """执行任意 SQL，自动区分 SELECT（返回结果集）与 DDL/DML（返回影响行数）.

    Args:
        engine: SQLAlchemy 引擎。
        sql: 原始 SQL 字符串（可含末尾分号）。
        read_only: 是否强制只读模式（True 时非只读语句抛 QueryError）。
            用于 viewer 角色限制：调用方传 ``read_only=True`` 拦截写操作。

    Returns:
        dict 含以下字段：

        - ``columns``: 列名列表（SELECT 时为结果集列名，DDL/DML 时为空列表）。
        - ``rows``: 行 dict 列表（SELECT 时为结果集；DDL/DML 时为空列表）。
        - ``rowcount``: 影响行数（SELECT 时为结果集行数；DDL/DML 时为底层 rowcount，无法获取时为 -1）。
        - ``elapsed_ms``: 执行耗时（毫秒，float）。
        - ``read_only``: 实际执行的语句是否为只读。

    Raises:
        QueryError: SQL 为空、``read_only=True`` 但语句非只读。
        SQLAlchemyError: 底层 SQL 执行失败（语法错误、约束冲突等）。
    """
    cleaned = _strip_sql(sql)
    is_read = _is_read_only(cleaned)
    if read_only and not is_read:
        raise QueryError("当前角色仅允许执行只读查询（SELECT/WITH/SHOW/DESCRIBE/EXPLAIN）")

    start = time.perf_counter()
    if is_read:
        # SELECT/WITH：使用 connect() 读取结果集，无显式事务
        with engine.connect() as conn:
            result = conn.execute(text(cleaned))
            columns_list: list[str] = list(result.keys()) if result.returns_rows else []
            rows_list: list[dict[str, Any]] = (
                [cast("dict[str, Any]", dict(row._mapping)) for row in result.fetchall()] if result.returns_rows else []
            )
        elapsed = (time.perf_counter() - start) * 1000
        return {
            "columns": columns_list,
            "rows": rows_list,
            "rowcount": len(rows_list),
            "elapsed_ms": round(elapsed, 3),
            "read_only": True,
        }

    # DDL/DML：使用 begin() 显式事务，统一 commit
    with engine.begin() as conn:
        result = conn.execute(text(cleaned))
        # rowcount：INSERT/UPDATE/DELETE 为影响行数；DDL 通常为 -1
        rowcount = result.rowcount
    elapsed = (time.perf_counter() - start) * 1000
    return {
        "columns": [],
        "rows": [],
        "rowcount": rowcount,
        "elapsed_ms": round(elapsed, 3),
        "read_only": False,
    }


def explain_sql(
    engine: Engine,
    sql: str,
    *,
    analyze: bool = False,
) -> dict[str, Any]:
    """获取 SQL 执行计划.

    多方言适配：

    - SQLite: ``EXPLAIN QUERY PLAN <sql>``（不支持 ANALYZE，``analyze`` 参数忽略）
    - PostgreSQL: ``EXPLAIN [ANALYZE] <sql>``
    - MySQL: ``EXPLAIN [ANALYZE] <sql>``（MySQL 8.0+ 支持 ANALYZE）

    Args:
        engine: SQLAlchemy 引擎。
        sql: 待分析的 SQL（通常为 SELECT；部分方言支持 DML 的 EXPLAIN）。
        analyze: 是否实际执行以获取真实执行统计（PG/MySQL 8.0+ 支持）。
            SQLite 忽略此参数。

    Returns:
        dict 含以下字段：

        - ``plan``: 执行计划文本行列表（每行一个字符串）。
        - ``rows``: 结构化行 dict 列表（保留原始列，便于前端表格展示）。
        - ``columns``: 结果列名列表。
        - ``analyze``: 实际是否启用 ANALYZE。
        - ``dialect``: 方言名。

    Raises:
        QueryError: SQL 为空、方言不支持 EXPLAIN。
        SQLAlchemyError: 底层执行失败。
    """
    cleaned = _strip_sql(sql)
    dialect = engine.dialect.name
    template = _EXPLAIN_TEMPLATES.get(dialect)
    if template is None:
        raise QueryError(f"方言 {dialect} 暂不支持 EXPLAIN")

    # SQLite 不支持 ANALYZE；其他方言按用户意图拼接
    analyze_keyword = "ANALYZE " if analyze and dialect != EngineType.SQLITE else ""
    explain_sql_str = template.format(sql=cleaned, analyze=analyze_keyword)

    with engine.connect() as conn:
        result = conn.execute(text(explain_sql_str))
        columns_list: list[str] = list(result.keys()) if result.returns_rows else []
        rows_raw: list[Any] = list(result.fetchall()) if result.returns_rows else []

    # SQLite EXPLAIN QUERY PLAN 返回 (id, parent, notused, detail) 四列；
    # PG/MySQL EXPLAIN 通常返回单列文本行（每行一段计划文本）。
    # 统一构造结构化行 + 文本行两种视图，便于前端展示。
    rows_list: list[dict[str, Any]] = [cast("dict[str, Any]", dict(row._mapping)) for row in rows_raw]
    plan_lines: list[str]
    if columns_list:
        plan_lines = [" | ".join(str(row.get(col, "")) for col in columns_list) for row in rows_list]
    else:  # pragma: no cover - 主流方言 EXPLAIN 均返回行集，防御性兜底
        plan_lines = []

    return {
        "plan": plan_lines,
        "rows": rows_list,
        "columns": columns_list,
        "analyze": bool(analyze_keyword),
        "dialect": dialect,
    }


# ============================================================
# P4-4 导入导出
# ============================================================


def _format_csv_value(val: Any) -> str:
    """格式化值为 CSV 单元格字符串.

    - ``None`` → 空字符串
    - ``bool`` → ``"1"``/``"0"``（避免 Python ``True``/``False`` 字面量）
    - ``datetime``/``date``/``time`` → ISO 格式字符串
    - ``bytes`` → UTF-8 解码（失败用 replace）
    - ``dict``/``list`` → JSON 字符串
    - 其他 → ``str(val)``
    """
    if val is None:
        return ""
    if isinstance(val, bool):
        return "1" if val else "0"
    if isinstance(val, (datetime, date, dtime)):
        return val.isoformat()
    if isinstance(val, bytes):
        return val.decode("utf-8", errors="replace")
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False, default=str)
    return str(val)


def _format_sql_value(val: Any, dialect: str) -> str:
    """格式化值为 SQL 字面量（用于生成 INSERT 脚本）.

    - ``None`` → ``NULL``
    - ``bool`` → ``1``/``0``（SQLite/PG/MySQL 整数布尔）
    - ``int``/``float`` → 数字字面量
    - ``datetime``/``date``/``time`` → ``'ISO 字符串'``
    - ``bytes`` → SQLite ``X'hex'``，其他 ``'utf-8 解码'``
    - 其他 → ``'转义后的字符串'``（单引号翻倍）
    """
    if val is None:
        return "NULL"
    if isinstance(val, bool):
        return "1" if val else "0"
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, (datetime, date, dtime)):
        return f"'{val.isoformat()}'"
    # bytes 与字符串统一转义处理：SQLite 用 X'hex'，其他方言解码为字符串
    if isinstance(val, bytes) and dialect == EngineType.SQLITE:
        return f"X'{val.hex()}'"
    if isinstance(val, bytes):
        val = val.decode("utf-8", errors="replace")
    s = str(val).replace("'", "''")
    return f"'{s}'"


def _format_excel_value(val: Any) -> Any:
    """格式化值为 Excel 单元格值（openpyxl 支持 str/int/float/datetime/bool/None）."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float, str, datetime, date, dtime)):
        return val
    if isinstance(val, bytes):
        return val.decode("utf-8", errors="replace")
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False, default=str)
    return str(val)


def iter_table_rows(
    engine: Engine,
    table_name: str,
    schema: str | None = None,
    batch_size: int = 1000,
) -> Iterator[dict[str, Any]]:
    """流式生成表行数据（生成器，避免大表 OOM）.

    使用 ``fetchmany(batch_size)`` 分批读取，PG/MySQL 启用服务端游标
    （``stream_results=True``）进一步降低内存占用。

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 表名。
        schema: Schema 名（SQLite 强制 None）。
        batch_size: 每批拉取行数，默认 1000。

    Yields:
        行数据 dict（键为列名）。

    Raises:
        QueryError: 表不存在或反射失败。
    """
    dialect = engine.dialect.name
    effective_schema = _resolve_schema(engine, schema)
    table_ref = _format_table_ref(table_name, effective_schema, dialect)
    select_sql = f"SELECT * FROM {table_ref}"

    # 先校验表存在（避免 SELECT 抛出底层异常时消息不友好）
    get_column_names(engine, table_name, schema)

    with engine.connect() as conn:
        # PG/MySQL 启用服务端游标；SQLite 忽略此选项
        exec_conn = conn.execution_options(stream_results=True) if dialect != EngineType.SQLITE else conn
        result = exec_conn.execute(text(select_sql))
        while True:
            batch = result.fetchmany(batch_size)
            if not batch:
                break
            for row in batch:
                yield cast("dict[str, Any]", dict(row._mapping))


def iter_filtered_table_rows(  # noqa: PLR0913, PLR0917
    engine: Engine,
    table_name: str,
    schema: str | None = None,
    columns: list[str] | None = None,
    filters: dict[str, dict[str, Any]] | None = None,
    batch_size: int = 1000,
) -> Iterator[dict[str, Any]]:
    """流式生成带列裁剪与筛选条件的表行数据（生成器）.

    与 :func:`iter_table_rows` 的差异：支持 ``columns`` 列裁剪与 ``filters``
    筛选条件，复用 :func:`query_table_rows` 的列名校验与 WHERE 构造逻辑，
    适用于数据集导出等需应用行级过滤/列级白名单的场景。

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 表名。
        schema: Schema 名（SQLite 强制 None）。
        columns: 显式指定的列名列表；None 或空表示所有列。
        filters: 筛选条件，格式 ``{列名: {"op": "eq/ne/gt/lt/ge/le/like/in", "val": ...}}``。
        batch_size: 每批拉取行数，默认 1000。

    Yields:
        行数据 dict（键为列名）。

    Raises:
        QueryError: 列名非法、操作符不支持、表不存在或反射失败。
    """
    filters = filters or {}
    dialect = engine.dialect.name
    effective_schema = _resolve_schema(engine, schema)
    table_ref = _format_table_ref(table_name, effective_schema, dialect)

    allowed = set(get_column_names(engine, table_name, schema))
    selected_columns = _validate_columns(allowed, columns)
    _validate_filter_columns(allowed, filters)

    if selected_columns:
        select_cols = ", ".join(_quote_ident(c, dialect) for c in selected_columns)
    else:
        select_cols = "*"

    where_sql, params = _build_where_clause(filters, dialect)
    where_clause = f" WHERE {where_sql}" if where_sql else ""
    select_sql = f"SELECT {select_cols} FROM {table_ref}{where_clause}"

    with engine.connect() as conn:
        exec_conn = conn.execution_options(stream_results=True) if dialect != EngineType.SQLITE else conn
        result = exec_conn.execute(text(select_sql), params)
        while True:
            batch = result.fetchmany(batch_size)
            if not batch:
                break
            for row in batch:
                yield cast("dict[str, Any]", dict(row._mapping))


def rows_to_csv(
    rows_iter: Iterator[dict[str, Any]],
    columns: list[str],
) -> Iterator[str]:
    """将行迭代器转为 CSV 文本块（生成器）.

    首块输出 UTF-8 BOM + 表头行（BOM 便于 Excel 识别中文），后续每行一个文本块。
    用 ``\\n`` 行尾，兼容主流工具。

    Args:
        rows_iter: 行 dict 迭代器。
        columns: 列名顺序（决定表头与每行字段顺序）。

    Yields:
        CSV 文本块（字符串）。
    """
    buf = io.StringIO()
    buf.write("\ufeff")  # UTF-8 BOM
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(columns)
    yield buf.getvalue()

    for row in rows_iter:
        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\n")
        writer.writerow([_format_csv_value(row.get(c)) for c in columns])
        yield buf.getvalue()


def rows_to_sql(
    rows_iter: Iterator[dict[str, Any]],
    columns: list[str],
    table_name: str,
    schema: str | None,
    dialect: str,
) -> Iterator[str]:
    """将行迭代器转为 INSERT SQL 语句（生成器）.

    每行生成一条 ``INSERT INTO "tbl" ("c1","c2") VALUES (1,'a');`` 语句。

    Args:
        rows_iter: 行 dict 迭代器。
        columns: 列名顺序。
        table_name: 目标表名。
        schema: Schema 名（SQLite 强制 None）。
        dialect: 方言名。

    Yields:
        INSERT 语句字符串（含末尾分号与换行）。
    """
    table_ref = _format_table_ref(table_name, schema, dialect)
    col_refs = ", ".join(_quote_ident(c, dialect) for c in columns)
    prefix = f"INSERT INTO {table_ref} ({col_refs}) VALUES "

    for row in rows_iter:
        values = [_format_sql_value(row.get(c), dialect) for c in columns]
        yield f"{prefix}({', '.join(values)});\n"


def export_excel(
    engine: Engine,
    table_name: str,
    schema: str | None = None,
) -> bytes:
    """导出表数据为 Excel xlsx 二进制（openpyxl write_only 流式写入）.

    用 ``write_only=True`` 模式逐行写入，避免大表内存爆炸。

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 表名。
        schema: Schema 名（SQLite 强制 None）。

    Returns:
        完整 xlsx 文件字节。

    Raises:
        QueryError: 表不存在或反射失败。
    """
    from openpyxl import Workbook

    columns = get_column_names(engine, table_name, schema)
    wb = Workbook(write_only=True)
    # Excel sheet 名最长 31 字符，截断避免 openpyxl 抛错
    ws = wb.create_sheet(title=table_name[:31])
    ws.append(columns)

    for row in iter_table_rows(engine, table_name, schema):
        ws.append([_format_excel_value(row.get(c)) for c in columns])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def iter_select_rows(
    engine: Engine,
    sql: str,
    *,
    read_only: bool = False,
    batch_size: int = 1000,
) -> tuple[list[str], Iterator[dict[str, Any]]]:
    """流式执行 SELECT 语句，返回 ``(列名列表, 行 dict 生成器)``.

    复用 ``_strip_sql``/``_is_read_only`` 进行前缀判断；强制 ``read_only=True``
    （调用方传入的 ``read_only`` 仅用于二次校验，导出场景禁止 DDL/DML）。
    PG/MySQL 启用 ``stream_results=True`` 服务端游标，``fetchmany(batch_size)`` 分批
    拉取，避免大结果集 OOM。

    Args:
        engine: SQLAlchemy 引擎。
        sql: 原始 SQL（可含末尾分号）。
        read_only: 是否要求语句本身为只读。导出场景调用方应传 ``True``，使
            非只读语句在执行前直接抛 ``QueryError``（防止导出走 DDL/DML）。
        batch_size: 每批拉取行数，默认 1000。

    Returns:
        ``(columns, rows_iter)``：``columns`` 为结果集列名列表；
        ``rows_iter`` 为行 dict 生成器（消费时维持连接打开）。

    Raises:
        QueryError: SQL 为空、语句非只读（``read_only=True`` 时）。
        SQLAlchemyError: SQL 执行失败（在函数返回前触发，因首行 ``next`` 已执行查询）。
    """
    cleaned = _strip_sql(sql)
    if read_only and not _is_read_only(cleaned):
        raise QueryError("导出仅允许只读查询（SELECT/WITH/SHOW/DESCRIBE/EXPLAIN）")

    dialect = engine.dialect.name
    state: dict[str, list[str]] = {"columns": []}

    def _stream() -> Iterator[dict[str, Any]]:
        with engine.connect() as conn:
            exec_conn = conn.execution_options(stream_results=True) if dialect != EngineType.SQLITE else conn
            result = exec_conn.execute(text(cleaned))
            state["columns"] = list(result.keys()) if result.returns_rows else []
            if not result.returns_rows:
                return
            while True:
                batch = result.fetchmany(batch_size)
                if not batch:
                    break
                for row in batch:
                    yield cast("dict[str, Any]", dict(row._mapping))

    gen = _stream()
    try:
        # 触发生成器执行查询并填充 state["columns"]；保留首行接回迭代器头部
        first_row = next(gen)
    except StopIteration:
        # 空结果集（无返回行）：columns 已填充，返回空迭代器
        return state["columns"], iter(())
    return state["columns"], itertools.chain([first_row], gen)


def rows_to_json(
    rows_iter: Iterator[dict[str, Any]],
    columns: list[str],
) -> Iterator[str]:
    """将行迭代器转为流式 JSON 数组文本块（生成器）.

    首块输出 ``[`` + 第一行 JSON（无前置逗号），后续每行输出 ``,\\n`` + 行 JSON，
    末块输出 ``]``。空结果集输出 ``[\\n]\\n``。

    Args:
        rows_iter: 行 dict 迭代器。
        columns: 列名顺序（仅作参考，行 dict 已含键值；保留参数与 ``rows_to_csv`` 签名对齐）。

    Yields:
        JSON 文本块（字符串，UTF-8 编码由调用方处理）。
    """
    del columns  # 行 dict 自带键，无需按 columns 重排

    first = True
    for row in rows_iter:
        if first:
            yield "[\n" + json.dumps(row, ensure_ascii=False, default=str)
            first = False
        else:
            yield ",\n" + json.dumps(row, ensure_ascii=False, default=str)
    if first:
        yield "[\n]\n"
    else:
        yield "\n]\n"


def export_sql_result_excel(engine: Engine, sql: str) -> bytes:
    """导出 SQL 结果集为 Excel xlsx 二进制（openpyxl write_only 流式写入）.

    复用 ``iter_select_rows`` 流式拉取数据，强制 ``read_only=True``（导出仅允许 SELECT）。
    sheet 名固定 ``query_result``（SQL 结果集无表名概念）。

    Args:
        engine: SQLAlchemy 引擎。
        sql: 原始 SELECT SQL（可含末尾分号）。

    Returns:
        完整 xlsx 文件字节。

    Raises:
        QueryError: SQL 为空或非只读。
        SQLAlchemyError: SQL 执行失败。
    """
    from openpyxl import Workbook

    columns, rows_iter = iter_select_rows(engine, sql, read_only=True)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="query_result")
    ws.append(columns)

    for row in rows_iter:
        ws.append([_format_excel_value(row.get(c)) for c in columns])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_csv_upload(file_obj: Any) -> tuple[list[str], Iterator[dict[str, Any]]]:
    """解析上传的 CSV 文件，返回 ``(列名列表, 行 dict 生成器)``.

    首行作为表头。支持 UTF-8 与 UTF-8 BOM。

    Args:
        file_obj: Django ``UploadedFile`` 或类似文件对象（支持 ``read()``）。

    Returns:
        ``(headers, rows_iter)``：``headers`` 为列名列表，``rows_iter`` 为行 dict 生成器。

    Raises:
        QueryError: 文件为空。
    """
    raw = file_obj.read()
    if isinstance(raw, bytes):
        # 去除 UTF-8 BOM
        if raw.startswith(b"\xef\xbb\xbf"):
            raw = raw[3:]
        text = raw.decode("utf-8", errors="replace")
    else:
        text = raw
        if text.startswith("\ufeff"):
            text = text[1:]

    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration as exc:
        raise QueryError("CSV 文件为空") from exc

    def _iter() -> Iterator[dict[str, Any]]:
        for row in reader:
            # 行长度可能与表头不一致（如末尾空字段），用 zip 容错
            yield dict(zip(headers, row, strict=False))

    return headers, _iter()


def parse_excel_upload(file_obj: Any) -> tuple[list[str], Iterator[dict[str, Any]]]:
    """解析上传的 Excel xlsx 文件，返回 ``(列名列表, 行 dict 生成器)``.

    首行作为表头。用 ``read_only=True`` 流式读取，避免大文件 OOM。

    Args:
        file_obj: Django ``UploadedFile`` 或类似文件对象。

    Returns:
        ``(headers, rows_iter)``。

    Raises:
        QueryError: 文件为空。
    """
    from openpyxl import load_workbook

    # Django UploadedFile 是类文件对象，openpyxl 可直接读取
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:  # pragma: no cover - openpyxl 加载有效 xlsx 必有 active sheet
        wb.close()
        raise QueryError("Excel 文件无工作表")
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_row = next(rows_iter)
    except StopIteration as exc:
        wb.close()
        raise QueryError("Excel 文件为空") from exc
    headers = [str(h) if h is not None else "" for h in headers_row]

    def _iter() -> Iterator[dict[str, Any]]:
        try:
            for row in rows_iter:
                yield dict(zip(headers, row, strict=False))
        finally:
            wb.close()

    return headers, _iter()


def import_rows(  # noqa: PLR0913, PLR0917
    engine: Engine,
    table_name: str,
    schema: str | None,
    columns: list[str],
    rows_iter: Iterator[dict[str, Any]],
    batch_size: int = 1000,
) -> dict[str, Any]:
    """事务批量插入行，任一失败回滚.

    所有行在单个 ``engine.begin()`` 事务内分批 ``executemany`` 插入。
    任一行违反约束（如 NOT NULL、类型不匹配、唯一约束）将抛 ``SQLAlchemyError``，
    事务自动回滚，已插入的批次一并撤销。

    Args:
        engine: SQLAlchemy 引擎。
        table_name: 目标表名。
        schema: Schema 名（SQLite 强制 None）。
        columns: 列名列表（与 ``rows_iter`` 中 dict 的键匹配；缺失列填 ``None`` 即 NULL）。
        rows_iter: 行 dict 迭代器。
        batch_size: 批量大小（用 ``executemany`` 一次插入），默认 1000。

    Returns:
        dict 含以下字段：

        - ``success_count``: 成功插入行数。
        - ``failed_count``: 始终 0（失败时抛错回滚，无部分插入）。
        - ``errors``: 空列表（同上）。

    Raises:
        QueryError: 列名列表为空、列名非法、表不存在。
        SQLAlchemyError: 底层插入失败（约束冲突等，事务已回滚）。
    """
    if not columns:
        raise QueryError("列名列表不能为空")

    allowed = set(get_column_names(engine, table_name, schema))
    for col in columns:
        if col not in allowed:
            raise QueryError(f"非法列名: {col}")

    dialect = engine.dialect.name
    effective_schema = _resolve_schema(engine, schema)
    table_ref = _format_table_ref(table_name, effective_schema, dialect)
    col_refs = ", ".join(_quote_ident(c, dialect) for c in columns)
    placeholders = ", ".join(f":v{i}" for i in range(len(columns)))
    insert_sql = f"INSERT INTO {table_ref} ({col_refs}) VALUES ({placeholders})"

    success_count = 0
    batch: list[dict[str, Any]] = []

    with engine.begin() as conn:
        for row in rows_iter:
            batch.append({f"v{i}": row.get(c) for i, c in enumerate(columns)})
            if len(batch) >= batch_size:
                conn.execute(text(insert_sql), batch)
                success_count += len(batch)
                batch = []
        if batch:
            conn.execute(text(insert_sql), batch)
            success_count += len(batch)

    return {
        "success_count": success_count,
        "failed_count": 0,
        "errors": [],
    }


__all__ = [
    "QueryError",
    "count_table_rows",
    "delete_row",
    "execute_sql",
    "explain_sql",
    "export_excel",
    "export_sql_result_excel",
    "get_column_names",
    "get_pk_columns",
    "get_row",
    "import_rows",
    "insert_row",
    "iter_filtered_table_rows",
    "iter_select_rows",
    "iter_table_rows",
    "parse_csv_upload",
    "parse_excel_upload",
    "query_table_rows",
    "rows_to_csv",
    "rows_to_json",
    "rows_to_sql",
    "update_row",
]
