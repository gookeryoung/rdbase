"""文件下载爬取 Spider.

从 source_url 下载文件（CSV/Excel/JSON），解析为行列表逐条 yield 为 dict。
不翻页（文件下载为单次请求）。

parse_config 结构::

    {
        "format": "csv",            // csv/excel/json（必填）
        "encoding": "utf-8",        // 可选，默认 utf-8（CSV/JSON）
        "delimiter": ",",           // 可选，CSV 分隔符（默认逗号）
        "sheet": "Sheet1",          // 可选，Excel 工作表名（默认第一个）
        "items_path": "$.data[*]"   // 可选，JSON 条目 JSONPath（默认视响应为列表）
    }
"""

from __future__ import annotations

import csv
import io
import json
import logging
from collections.abc import Iterator
from typing import Any

from scrapy.http import Request, Response  # type: ignore[import-not-found]

from apps.ingest.spiders.base import BaseIngestSpider

logger = logging.getLogger(__name__)


class FileIngestSpider(BaseIngestSpider):
    """文件下载爬取 Spider.

    从 source_url 下载文件，按 format 配置解析为行列表，
    逐行 yield 为 dict 供 pipeline 处理。
    """

    name = "ingest_file"

    def __init__(self, **kwargs: Any) -> None:
        super().__init__(**kwargs)
        if self.source_url:
            self.start_urls = [self.source_url]

    def start_requests(self) -> Iterator[Request]:  # type: ignore[missing-override-decorator, override]
        """发起文件下载请求."""
        for url in self.start_urls:
            yield Request(url, headers=self.headers or None, callback=self.parse, dont_filter=True)

    def parse(self, response: Response, **_kwargs: Any) -> Iterator[Any]:  # type: ignore[missing-override-decorator, override]
        """按文件格式解析响应体为行列表.

        Args:
            response: Scrapy 下载器返回的响应对象（body 为文件字节）。
        """
        fmt = str(self.parse_config.get("format", "")).lower()
        body: bytes = response.body

        if fmt == "csv":
            yield from self._parse_csv(body)
        elif fmt == "excel":
            yield from self._parse_excel(body)
        elif fmt == "json":
            yield from self._parse_json(body)
        else:
            logger.error("不支持的文件格式: %s（支持 csv/excel/json）", fmt)

    def _parse_csv(self, body: bytes) -> Iterator[dict[str, Any]]:
        """解析 CSV 文件为行字典."""
        encoding = str(self.parse_config.get("encoding", "utf-8"))
        delimiter = str(self.parse_config.get("delimiter", ","))
        text = body.decode(encoding, errors="replace")
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        for row in reader:
            yield dict(row)

    def _parse_excel(self, body: bytes) -> Iterator[dict[str, Any]]:
        """解析 Excel 文件为行字典."""
        from openpyxl import load_workbook  # type: ignore[import-not-found]

        sheet_name = self.parse_config.get("sheet")
        wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb.active
            rows = ws.iter_rows(values_only=True)
            headers: list[str] | None = None
            for row in rows:
                if headers is None:
                    headers = [str(c) if c is not None else "" for c in row]
                    continue
                yield {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        finally:
            wb.close()

    def _parse_json(self, body: bytes) -> Iterator[dict[str, Any]]:
        """解析 JSON 文件为行字典（支持 JSONPath 定位条目数组）."""
        encoding = str(self.parse_config.get("encoding", "utf-8"))
        text = body.decode(encoding, errors="replace")

        try:
            data = json.loads(text)
        except (json.JSONDecodeError, ValueError) as exc:
            logger.error("JSON 文件解析失败: %s", exc)
            return

        items_path = self.parse_config.get("items_path")
        if not items_path:
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict):
                        yield item
            elif isinstance(data, dict):
                yield data
            return

        try:
            from jsonpath_ng.ext import parse as jsonpath_parse  # type: ignore[import-not-found]

            expr = jsonpath_parse(str(items_path))
        except Exception as exc:
            logger.error("JSONPath 解析失败: %s, error: %s", items_path, exc)
            return

        for match in expr.find(data):
            value = match.value
            if isinstance(value, dict):
                yield value
            elif isinstance(value, list):
                for sub in value:
                    if isinstance(sub, dict):
                        yield sub


__all__ = ["FileIngestSpider"]
