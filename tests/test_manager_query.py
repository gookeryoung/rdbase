"""manager 数据查询与 CRUD 模块单元测试.

使用 SQLite 内存库 + StaticPool（单连接，确保表跨连接可见）建表后验证查询与 CRUD 结果。
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, cast

import pytest
from apps.manager.query import (
    QueryError,
    _build_where_clause,
    _format_csv_value,
    _format_excel_value,
    _format_sql_value,
    _format_table_ref,
    _is_read_only,
    _quote_ident,
    _resolve_schema,
    _strip_sql,
    count_table_rows,
    delete_row,
    execute_sql,
    explain_sql,
    export_excel,
    get_column_names,
    get_pk_columns,
    get_row,
    import_rows,
    insert_row,
    iter_table_rows,
    parse_csv_upload,
    parse_excel_upload,
    query_table_rows,
    rows_to_csv,
    rows_to_sql,
    update_row,
)
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine
from sqlalchemy.pool import StaticPool


def _make_memory_engine() -> Engine:
    """构造 SQLite 内存库引擎（StaticPool 单连接，确保建表后跨连接可见）."""
    return cast(
        Engine,
        create_engine(
            "sqlite:///:memory:",
            poolclass=StaticPool,
            future=True,
        ),
    )


def _setup_tables(engine: Engine) -> None:
    """在引擎中创建测试表并插入数据."""
    with engine.begin() as conn:
        conn.execute(
            text(
                "CREATE TABLE users ("
                "id INTEGER PRIMARY KEY AUTOINCREMENT, "
                "name VARCHAR(50) NOT NULL, "
                "email VARCHAR(100), "
                "age INTEGER DEFAULT 0"
                ")"
            )
        )
        conn.execute(
            text(
                "INSERT INTO users (id, name, email, age) VALUES "
                "(1, 'Alice', 'alice@example.com', 30), "
                "(2, 'Bob', 'bob@example.com', 25), "
                "(3, 'Charlie', 'charlie@example.com', 35), "
                "(4, 'David', 'david@example.com', 28), "
                "(5, 'Eve', 'eve@example.com', 22)"
            )
        )


# ---------- 标识符引用 ----------


def test_quote_ident_mysql_uses_backticks() -> None:
    """MySQL 方言应使用反引号引用标识符."""
    assert _quote_ident("name", "mysql") == "`name`"


def test_quote_ident_non_mysql_uses_double_quotes() -> None:
    """非 MySQL 方言应使用双引号引用标识符."""
    assert _quote_ident("name", "postgresql") == '"name"'
    assert _quote_ident("name", "sqlite") == '"name"'


def test_format_table_ref_with_schema() -> None:
    """非 SQLite 方言应包含 schema 前缀."""
    assert _format_table_ref("users", "public", "postgresql") == '"public"."users"'
    assert _format_table_ref("users", "public", "mysql") == "`public`.`users`"


def test_format_table_ref_sqlite_ignores_schema() -> None:
    """SQLite 方言应忽略 schema."""
    assert _format_table_ref("users", "main", "sqlite") == '"users"'
    assert _format_table_ref("users", None, "sqlite") == '"users"'


# ---------- _resolve_schema ----------


def test_resolve_schema_sqlite_returns_none() -> None:
    """SQLite 应强制返回 None."""
    engine = _make_memory_engine()
    try:
        assert _resolve_schema(engine, "main") is None
    finally:
        engine.dispose()


class _FakeDialect:
    def __init__(self, name: str) -> None:
        self.name = name


class _FakeEngine:
    def __init__(self, dialect_name: str) -> None:
        self.dialect = _FakeDialect(dialect_name)

    def connect(self) -> Any:  # pragma: no cover - 未在 resolve_schema 中调用
        raise AssertionError("不应调用 connect")


def test_resolve_schema_non_sqlite_returns_input() -> None:
    """非 SQLite 方言应原样返回 schema（用 FakeEngine 覆盖分支）."""
    engine = _FakeEngine("mysql")
    assert _resolve_schema(cast(Engine, engine), "public") == "public"


# ---------- get_column_names ----------


def test_get_column_names_returns_all() -> None:
    """应返回表的所有列名."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        cols = get_column_names(engine, "users", schema=None)
        assert cols == ["id", "name", "email", "age"]
    finally:
        engine.dispose()


def test_get_column_names_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            get_column_names(engine, "nonexistent", schema=None)
    finally:
        engine.dispose()


# ---------- _build_where_clause ----------


def test_build_where_clause_empty_filters() -> None:
    """空 filters 应返回空字符串与空参数."""
    sql, params = _build_where_clause({}, "sqlite")
    assert sql == ""
    assert params == {}


def test_build_where_clause_eq() -> None:
    """eq 操作符应生成 = 占位符."""
    sql, params = _build_where_clause({"name": {"op": "eq", "val": "Alice"}}, "sqlite")
    assert sql == '"name" = :f0'
    assert params == {"f0": "Alice"}


def test_build_where_clause_in() -> None:
    """in 操作符应展开多个占位符."""
    sql, params = _build_where_clause({"id": {"op": "in", "val": [1, 2, 3]}}, "sqlite")
    assert sql == '"id" IN (:f0_0, :f0_1, :f0_2)'
    assert params == {"f0_0": 1, "f0_1": 2, "f0_2": 3}


def test_build_where_clause_in_empty_list_raises() -> None:
    """in 操作符空列表应抛 QueryError."""
    with pytest.raises(QueryError):
        _build_where_clause({"id": {"op": "in", "val": []}}, "sqlite")


def test_build_where_clause_multiple_filters() -> None:
    """多个筛选条件应用 AND 连接."""
    sql, params = _build_where_clause(
        {"name": {"op": "like", "val": "A%"}, "age": {"op": "gt", "val": 18}},
        "sqlite",
    )
    assert sql == '"name" LIKE :f0 AND "age" > :f1'
    assert params == {"f0": "A%", "f1": 18}


def test_build_where_clause_all_comparators() -> None:
    """应支持全部比较操作符."""
    ops = ["eq", "ne", "gt", "lt", "ge", "le", "like"]
    expected_symbols = ["=", "<>", ">", "<", ">=", "<=", "LIKE"]
    for op, symbol in zip(ops, expected_symbols, strict=True):
        sql, _ = _build_where_clause({"col": {"op": op, "val": 1}}, "sqlite")
        assert sql == f'"col" {symbol} :f0'


# ---------- query_table_rows ----------


def test_query_table_rows_default() -> None:
    """默认应返回所有行、所有列."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, total = query_table_rows(engine, "users", schema=None)
        assert total == 5
        assert len(rows) == 5
        assert set(rows[0].keys()) == {"id", "name", "email", "age"}
    finally:
        engine.dispose()


def test_query_table_rows_pagination() -> None:
    """分页应正确返回对应页的行."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, total = query_table_rows(engine, "users", schema=None, page=2, page_size=2)
        assert total == 5
        assert len(rows) == 2
        # 第二页应返回 id=3, 4（按插入顺序）
        assert [r["id"] for r in rows] == [3, 4]
    finally:
        engine.dispose()


def test_query_table_rows_pagination_beyond_end() -> None:
    """分页超出末尾应返回空列表但 total 不变."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, total = query_table_rows(engine, "users", schema=None, page=10, page_size=2)
        assert total == 5
        assert rows == []
    finally:
        engine.dispose()


def test_query_table_rows_order_by_asc() -> None:
    """应支持升序排序."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, _ = query_table_rows(
            engine,
            "users",
            schema=None,
            order_by="age",
            order_dir="asc",
        )
        ages = [r["age"] for r in rows]
        assert ages == sorted(ages)
        assert ages[0] == 22  # Eve 最年轻
    finally:
        engine.dispose()


def test_query_table_rows_order_by_desc() -> None:
    """应支持降序排序."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, _ = query_table_rows(
            engine,
            "users",
            schema=None,
            order_by="age",
            order_dir="desc",
        )
        ages = [r["age"] for r in rows]
        assert ages == sorted(ages, reverse=True)
        assert ages[0] == 35  # Charlie 最年长
    finally:
        engine.dispose()


def test_query_table_rows_columns_subset() -> None:
    """应支持显式指定返回的列."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, _ = query_table_rows(
            engine,
            "users",
            schema=None,
            columns=["id", "name"],
        )
        assert len(rows) == 5
        assert set(rows[0].keys()) == {"id", "name"}
    finally:
        engine.dispose()


def test_query_table_rows_filter_eq() -> None:
    """应支持 eq 筛选."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, total = query_table_rows(
            engine,
            "users",
            schema=None,
            filters={"name": {"op": "eq", "val": "Alice"}},
        )
        assert total == 1
        assert len(rows) == 1
        assert rows[0]["name"] == "Alice"
    finally:
        engine.dispose()


def test_query_table_rows_filter_like() -> None:
    """应支持 like 模糊筛选."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, total = query_table_rows(
            engine,
            "users",
            schema=None,
            filters={"email": {"op": "like", "val": "%@example.com"}},
        )
        assert total == 5
        assert len(rows) == 5
    finally:
        engine.dispose()


def test_query_table_rows_filter_in() -> None:
    """应支持 in 筛选."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, total = query_table_rows(
            engine,
            "users",
            schema=None,
            filters={"id": {"op": "in", "val": [1, 3, 5]}},
        )
        assert total == 3
        assert {r["id"] for r in rows} == {1, 3, 5}
    finally:
        engine.dispose()


def test_query_table_rows_filter_combined() -> None:
    """应支持多列组合筛选."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, total = query_table_rows(
            engine,
            "users",
            schema=None,
            filters={
                "age": {"op": "gt", "val": 25},
                "name": {"op": "like", "val": "A%"},
            },
        )
        assert total == 1
        assert rows[0]["name"] == "Alice"
    finally:
        engine.dispose()


def test_query_table_rows_invalid_column_raises() -> None:
    """非法列名应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            query_table_rows(engine, "users", schema=None, columns=["nonexistent"])
    finally:
        engine.dispose()


def test_query_table_rows_invalid_order_by_raises() -> None:
    """非法排序字段应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            query_table_rows(engine, "users", schema=None, order_by="nonexistent")
    finally:
        engine.dispose()


def test_query_table_rows_invalid_filter_column_raises() -> None:
    """非法筛选列名应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            query_table_rows(
                engine,
                "users",
                schema=None,
                filters={"nonexistent": {"op": "eq", "val": 1}},
            )
    finally:
        engine.dispose()


def test_query_table_rows_invalid_op_raises() -> None:
    """非法操作符应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            query_table_rows(
                engine,
                "users",
                schema=None,
                filters={"name": {"op": "contains", "val": "A"}},
            )
    finally:
        engine.dispose()


def test_query_table_rows_invalid_page_raises() -> None:
    """page < 1 应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            query_table_rows(engine, "users", schema=None, page=0)
    finally:
        engine.dispose()


def test_query_table_rows_invalid_page_size_raises() -> None:
    """page_size < 1 应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            query_table_rows(engine, "users", schema=None, page_size=0)
    finally:
        engine.dispose()


def test_query_table_rows_invalid_order_dir_raises() -> None:
    """非法 order_dir 应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            query_table_rows(engine, "users", schema=None, order_dir="random")
    finally:
        engine.dispose()


def test_query_table_rows_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            query_table_rows(engine, "nonexistent", schema=None)
    finally:
        engine.dispose()


def test_query_table_rows_schema_ignored_for_sqlite() -> None:
    """SQLite 应忽略 schema 参数."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows, total = query_table_rows(engine, "users", schema="main")
        assert total == 5
        assert len(rows) == 5
    finally:
        engine.dispose()


# ---------- count_table_rows ----------


def test_count_table_rows_default() -> None:
    """应统计全表行数."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        assert count_table_rows(engine, "users", schema=None) == 5
    finally:
        engine.dispose()


def test_count_table_rows_with_filter() -> None:
    """应统计满足筛选的行数."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        assert (
            count_table_rows(
                engine,
                "users",
                schema=None,
                filters={"age": {"op": "gt", "val": 25}},
            )
            == 3
        )
    finally:
        engine.dispose()


def test_count_table_rows_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            count_table_rows(engine, "nonexistent", schema=None)
    finally:
        engine.dispose()


def test_count_table_rows_invalid_filter_column_raises() -> None:
    """非法筛选列名应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            count_table_rows(
                engine,
                "users",
                schema=None,
                filters={"nonexistent": {"op": "eq", "val": 1}},
            )
    finally:
        engine.dispose()


# ---------- P4-2 行 CRUD ----------


def _setup_multi_pk_table(engine: Engine) -> None:
    """创建多列主键表并插入数据."""
    with engine.begin() as conn:
        conn.execute(
            text(
                "CREATE TABLE composite (a INTEGER NOT NULL, b INTEGER NOT NULL, name VARCHAR(50), PRIMARY KEY (a, b))"
            )
        )
        conn.execute(text("INSERT INTO composite (a, b, name) VALUES (1, 100, 'first'), (2, 200, 'second')"))


def _setup_no_pk_table(engine: Engine) -> None:
    """创建无主键表."""
    with engine.begin() as conn:
        conn.execute(text("CREATE TABLE no_pk (label VARCHAR(50))"))
        conn.execute(text("INSERT INTO no_pk (label) VALUES ('x'), ('y')"))


# ---------- get_pk_columns ----------


def test_get_pk_columns_single() -> None:
    """单列主键表应返回 [id]."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        assert get_pk_columns(engine, "users", schema=None) == ["id"]
    finally:
        engine.dispose()


def test_get_pk_columns_composite() -> None:
    """多列主键表应返回 [a, b]."""
    engine = _make_memory_engine()
    try:
        _setup_multi_pk_table(engine)
        assert get_pk_columns(engine, "composite", schema=None) == ["a", "b"]
    finally:
        engine.dispose()


def test_get_pk_columns_no_pk() -> None:
    """无主键表应返回空列表."""
    engine = _make_memory_engine()
    try:
        _setup_no_pk_table(engine)
        assert get_pk_columns(engine, "no_pk", schema=None) == []
    finally:
        engine.dispose()


def test_get_pk_columns_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            get_pk_columns(engine, "nonexistent", schema=None)
    finally:
        engine.dispose()


# ---------- insert_row ----------


def test_insert_row_with_autoincrement_pk() -> None:
    """自增主键场景应回填主键并返回完整行."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        row = insert_row(
            engine,
            "users",
            schema=None,
            values={"name": "Frank", "email": "frank@example.com", "age": 40},
        )
        # 主键 id 应被回填（>5，因为已有 5 行）
        assert row["id"] > 5
        assert row["name"] == "Frank"
        assert row["email"] == "frank@example.com"
        assert row["age"] == 40
        # 数据应已写入
        _rows, total = query_table_rows(engine, "users", schema=None)
        assert total == 6
    finally:
        engine.dispose()


def test_insert_row_with_explicit_pk() -> None:
    """显式提供主键值时应使用该值."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        row = insert_row(
            engine,
            "users",
            schema=None,
            values={"id": 100, "name": "Greg", "email": None, "age": 50},
        )
        assert row["id"] == 100
        assert row["name"] == "Greg"
    finally:
        engine.dispose()


def test_insert_row_composite_pk_explicit() -> None:
    """多列主键表需显式提供全部主键列."""
    engine = _make_memory_engine()
    try:
        _setup_multi_pk_table(engine)
        row = insert_row(
            engine,
            "composite",
            schema=None,
            values={"a": 3, "b": 300, "name": "third"},
        )
        assert row["a"] == 3
        assert row["b"] == 300
        assert row["name"] == "third"
    finally:
        engine.dispose()


def test_insert_row_empty_values_raises() -> None:
    """values 为空应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            insert_row(engine, "users", schema=None, values={})
    finally:
        engine.dispose()


def test_insert_row_invalid_column_raises() -> None:
    """非法列名应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            insert_row(
                engine,
                "users",
                schema=None,
                values={"nonexistent": "x"},
            )
    finally:
        engine.dispose()


def test_insert_row_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            insert_row(
                engine,
                "nonexistent",
                schema=None,
                values={"a": 1},
            )
    finally:
        engine.dispose()


def test_insert_row_no_pk_returns_values() -> None:
    """无主键表应直接返回传入的 values."""
    engine = _make_memory_engine()
    try:
        _setup_no_pk_table(engine)
        row = insert_row(
            engine,
            "no_pk",
            schema=None,
            values={"label": "z"},
        )
        assert row == {"label": "z"}
        # 数据应已写入
        _rows, total = query_table_rows(engine, "no_pk", schema=None)
        assert total == 3
    finally:
        engine.dispose()


def test_insert_row_composite_pk_missing_raises() -> None:
    """多列自增主键场景：未提供全部主键列且无法用 lastrowid 回填多列应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_multi_pk_table(engine)
        with pytest.raises(QueryError):
            # 只提供 a，缺 b，且 b 非自增（多列主键 lastrowid 无法定位单列）
            insert_row(
                engine,
                "composite",
                schema=None,
                values={"a": 5, "name": "missing"},
            )
    finally:
        engine.dispose()


# ---------- update_row ----------


def test_update_row_success() -> None:
    """成功更新单行并返回更新后的行."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        row = update_row(
            engine,
            "users",
            schema=None,
            pk={"id": 1},
            values={"name": "Alice2", "age": 31},
        )
        assert row["id"] == 1
        assert row["name"] == "Alice2"
        assert row["age"] == 31
        # email 应保持不变
        assert row["email"] == "alice@example.com"
    finally:
        engine.dispose()


def test_update_row_composite_pk() -> None:
    """多列主键表更新应支持."""
    engine = _make_memory_engine()
    try:
        _setup_multi_pk_table(engine)
        row = update_row(
            engine,
            "composite",
            schema=None,
            pk={"a": 1, "b": 100},
            values={"name": "updated"},
        )
        assert row["a"] == 1
        assert row["b"] == 100
        assert row["name"] == "updated"
    finally:
        engine.dispose()


def test_update_row_not_exists_raises() -> None:
    """行不存在应抛 QueryError（乐观锁 0 行）."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="不存在"):
            update_row(
                engine,
                "users",
                schema=None,
                pk={"id": 9999},
                values={"name": "ghost"},
            )
    finally:
        engine.dispose()


def test_update_row_empty_pk_raises() -> None:
    """主键为空应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            update_row(
                engine,
                "users",
                schema=None,
                pk={},
                values={"name": "x"},
            )
    finally:
        engine.dispose()


def test_update_row_empty_values_raises() -> None:
    """values 为空应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            update_row(
                engine,
                "users",
                schema=None,
                pk={"id": 1},
                values={},
            )
    finally:
        engine.dispose()


def test_update_row_invalid_pk_column_raises() -> None:
    """非法主键列名应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            update_row(
                engine,
                "users",
                schema=None,
                pk={"nonexistent": 1},
                values={"name": "x"},
            )
    finally:
        engine.dispose()


def test_update_row_invalid_value_column_raises() -> None:
    """非法更新列名应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            update_row(
                engine,
                "users",
                schema=None,
                pk={"id": 1},
                values={"nonexistent": "x"},
            )
    finally:
        engine.dispose()


def test_update_row_pk_in_values_raises() -> None:
    """主键列出现在 values 中应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="主键列"):
            update_row(
                engine,
                "users",
                schema=None,
                pk={"id": 1},
                values={"id": 2, "name": "x"},
            )
    finally:
        engine.dispose()


def test_update_row_no_pk_table_raises() -> None:
    """无主键表更新应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_no_pk_table(engine)
        with pytest.raises(QueryError, match="无主键"):
            update_row(
                engine,
                "no_pk",
                schema=None,
                pk={"label": "x"},
                values={"label": "y"},
            )
    finally:
        engine.dispose()


def test_update_row_pk_mismatch_raises() -> None:
    """主键列不匹配应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="不匹配"):
            update_row(
                engine,
                "users",
                schema=None,
                pk={"id": 1, "name": "Alice"},
                values={"age": 99},
            )
    finally:
        engine.dispose()


def test_update_row_composite_pk_partial_raises() -> None:
    """多列主键只提供部分应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_multi_pk_table(engine)
        with pytest.raises(QueryError, match="不匹配"):
            update_row(
                engine,
                "composite",
                schema=None,
                pk={"a": 1},
                values={"name": "x"},
            )
    finally:
        engine.dispose()


def test_update_row_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            update_row(
                engine,
                "nonexistent",
                schema=None,
                pk={"id": 1},
                values={"name": "x"},
            )
    finally:
        engine.dispose()


# ---------- delete_row ----------


def test_delete_row_success() -> None:
    """成功删除单行."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        delete_row(engine, "users", schema=None, pk={"id": 1})
        rows, total = query_table_rows(engine, "users", schema=None)
        assert total == 4
        assert all(r["id"] != 1 for r in rows)
    finally:
        engine.dispose()


def test_delete_row_composite_pk() -> None:
    """多列主键表删除应支持."""
    engine = _make_memory_engine()
    try:
        _setup_multi_pk_table(engine)
        delete_row(engine, "composite", schema=None, pk={"a": 1, "b": 100})
        rows, total = query_table_rows(engine, "composite", schema=None)
        assert total == 1
        assert rows[0]["a"] == 2
    finally:
        engine.dispose()


def test_delete_row_not_exists_raises() -> None:
    """行不存在应抛 QueryError（乐观锁 0 行）."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="不存在"):
            delete_row(engine, "users", schema=None, pk={"id": 9999})
    finally:
        engine.dispose()


def test_delete_row_empty_pk_raises() -> None:
    """主键为空应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            delete_row(engine, "users", schema=None, pk={})
    finally:
        engine.dispose()


def test_delete_row_invalid_pk_column_raises() -> None:
    """非法主键列名应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            delete_row(engine, "users", schema=None, pk={"nonexistent": 1})
    finally:
        engine.dispose()


def test_delete_row_no_pk_table_raises() -> None:
    """无主键表删除应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_no_pk_table(engine)
        with pytest.raises(QueryError, match="无主键"):
            delete_row(engine, "no_pk", schema=None, pk={"label": "x"})
    finally:
        engine.dispose()


def test_delete_row_pk_mismatch_raises() -> None:
    """主键列不匹配应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="不匹配"):
            delete_row(
                engine,
                "users",
                schema=None,
                pk={"id": 1, "name": "Alice"},
            )
    finally:
        engine.dispose()


def test_delete_row_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            delete_row(engine, "nonexistent", schema=None, pk={"id": 1})
    finally:
        engine.dispose()


# ---------- get_row ----------


def test_get_row_success() -> None:
    """按主键查单行应返回完整行."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        row = get_row(engine, "users", schema=None, pk={"id": 2})
        assert row is not None
        assert row["id"] == 2
        assert row["name"] == "Bob"
    finally:
        engine.dispose()


def test_get_row_not_exists_returns_none() -> None:
    """行不存在应返回 None."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        assert get_row(engine, "users", schema=None, pk={"id": 9999}) is None
    finally:
        engine.dispose()


def test_get_row_composite_pk() -> None:
    """多列主键查单行应支持."""
    engine = _make_memory_engine()
    try:
        _setup_multi_pk_table(engine)
        row = get_row(engine, "composite", schema=None, pk={"a": 2, "b": 200})
        assert row is not None
        assert row["name"] == "second"
    finally:
        engine.dispose()


def test_get_row_empty_pk_raises() -> None:
    """主键为空应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            get_row(engine, "users", schema=None, pk={})
    finally:
        engine.dispose()


def test_get_row_invalid_pk_column_raises() -> None:
    """非法主键列名应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError):
            get_row(engine, "users", schema=None, pk={"nonexistent": 1})
    finally:
        engine.dispose()


def test_get_row_no_pk_table_raises() -> None:
    """无主键表查询应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_no_pk_table(engine)
        with pytest.raises(QueryError, match="无主键"):
            get_row(engine, "no_pk", schema=None, pk={"label": "x"})
    finally:
        engine.dispose()


def test_get_row_pk_mismatch_raises() -> None:
    """主键列不匹配应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="不匹配"):
            get_row(engine, "users", schema=None, pk={"id": 1, "name": "Alice"})
    finally:
        engine.dispose()


def test_get_row_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            get_row(engine, "nonexistent", schema=None, pk={"id": 1})
    finally:
        engine.dispose()


# ---------- insert_row / update_row 异常分支补充（覆盖 query.py:402/408/481）----------


def test_insert_row_lastrowid_none_raises(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """lastrowid 为 None（dialect 不支持自增回填）应抛 QueryError（覆盖 query.py:402）.

    通过在 query 模块命名空间注入 ``getattr`` 替身，使 ``result.lastrowid`` 解析为 None，
    触发单列自增主键无法回填的异常分支。
    """
    real_getattr = getattr

    def mock_getattr(obj: Any, name: str, *default: Any) -> Any:
        if name == "lastrowid":
            return None
        return real_getattr(obj, name, *default)

    monkeypatch.setattr("apps.manager.query.getattr", mock_getattr, raising=False)
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="无法获取自增主键"):
            insert_row(
                engine,
                "users",
                schema=None,
                values={"name": "Ghost", "email": "g@e.com", "age": 1},
            )
    finally:
        engine.dispose()


def test_insert_row_post_select_none_raises(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """插入后回查返回 None 应抛 QueryError（覆盖 query.py:408）.

    模拟插入成功但同一事务内按主键反查返回 None（并发删除等极端场景）。
    """
    monkeypatch.setattr(
        "apps.manager.query._select_row_by_pk",
        lambda *_args, **_kw: None,
    )
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="插入后回查失败"):
            insert_row(
                engine,
                "users",
                schema=None,
                values={"name": "Ghost", "email": "g@e.com", "age": 1},
            )
    finally:
        engine.dispose()


def test_update_row_post_select_none_raises(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """更新后回查返回 None 应抛 QueryError（覆盖 query.py:481）.

    模拟更新成功（rowcount=1）但同一事务内按主键反查返回 None（并发删除等极端场景）。
    """
    monkeypatch.setattr(
        "apps.manager.query._select_row_by_pk",
        lambda *_args, **_kw: None,
    )
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="更新后回查失败"):
            update_row(
                engine,
                "users",
                schema=None,
                pk={"id": 1},
                values={"name": "Alice2"},
            )
    finally:
        engine.dispose()


# ============================================================
# P4-3 SQL 查询控制台 - _strip_sql / _is_read_only
# ============================================================


def test_strip_sql_strips_whitespace_and_semicolon() -> None:
    """应去除首尾空白与末尾分号."""
    assert _strip_sql("  SELECT 1;  ") == "SELECT 1"
    assert _strip_sql("SELECT 1") == "SELECT 1"
    assert _strip_sql("  SELECT 1; ") == "SELECT 1"


def test_strip_sql_empty_raises() -> None:
    """空 SQL 应抛 QueryError."""
    with pytest.raises(QueryError, match="不能为空"):
        _strip_sql("")
    with pytest.raises(QueryError, match="不能为空"):
        _strip_sql("   ")
    with pytest.raises(QueryError, match="不能为空"):
        _strip_sql(";")
    with pytest.raises(QueryError, match="不能为空"):
        _strip_sql("  ;  ")


def test_is_read_only_select() -> None:
    """SELECT/WITH 应识别为只读（输入须已去首尾空白）."""
    assert _is_read_only("SELECT * FROM users") is True
    assert _is_read_only("select * from users") is True
    assert _is_read_only("SELECT 1") is True
    assert _is_read_only("WITH t AS (SELECT 1) SELECT * FROM t") is True


def test_is_read_only_show_describe_explain() -> None:
    """SHOW/DESCRIBE/DESC/EXPLAIN 应识别为只读."""
    assert _is_read_only("SHOW TABLES") is True
    assert _is_read_only("DESCRIBE users") is True
    assert _is_read_only("DESC users") is True
    assert _is_read_only("EXPLAIN SELECT 1") is True


def test_is_read_only_dml_ddl() -> None:
    """DML/DDL 应识别为非只读."""
    assert _is_read_only("INSERT INTO users VALUES (1)") is False
    assert _is_read_only("UPDATE users SET name='x'") is False
    assert _is_read_only("DELETE FROM users") is False
    assert _is_read_only("CREATE TABLE t (id INT)") is False
    assert _is_read_only("DROP TABLE t") is False
    assert _is_read_only("ALTER TABLE t ADD COLUMN c INT") is False


# ============================================================
# P4-3 SQL 查询控制台 - execute_sql
# ============================================================


def test_execute_sql_select_returns_resultset() -> None:
    """SELECT 应返回结果集（columns/rows/rowcount/elapsed_ms/read_only）."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = execute_sql(engine, "SELECT id, name FROM users WHERE id = 1")
        assert result["read_only"] is True
        assert result["columns"] == ["id", "name"]
        assert len(result["rows"]) == 1
        assert result["rows"][0]["name"] == "Alice"
        assert result["rowcount"] == 1
        assert result["elapsed_ms"] >= 0
    finally:
        engine.dispose()


def test_execute_sql_select_trailing_semicolon() -> None:
    """末尾分号应被正确处理."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = execute_sql(engine, "SELECT COUNT(*) AS cnt FROM users;")
        assert result["columns"] == ["cnt"]
        assert result["rows"][0]["cnt"] == 5
    finally:
        engine.dispose()


def test_execute_sql_select_star() -> None:
    """SELECT * 应返回所有列."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = execute_sql(engine, "SELECT * FROM users")
        assert set(result["columns"]) == {"id", "name", "email", "age"}
        assert len(result["rows"]) == 5
    finally:
        engine.dispose()


def test_execute_sql_insert_returns_rowcount() -> None:
    """INSERT 应返回影响行数，read_only=False."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = execute_sql(
            engine,
            "INSERT INTO users (name, email, age) VALUES ('Frank', 'f@e.com', 40)",
        )
        assert result["read_only"] is False
        assert result["columns"] == []
        assert result["rows"] == []
        assert result["rowcount"] == 1
    finally:
        engine.dispose()


def test_execute_sql_update_returns_rowcount() -> None:
    """UPDATE 应返回影响行数."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = execute_sql(
            engine,
            "UPDATE users SET age = 99 WHERE id = 1",
        )
        assert result["read_only"] is False
        assert result["rowcount"] == 1
    finally:
        engine.dispose()


def test_execute_sql_delete_returns_rowcount() -> None:
    """DELETE 应返回影响行数."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = execute_sql(engine, "DELETE FROM users WHERE id = 1")
        assert result["read_only"] is False
        assert result["rowcount"] == 1
    finally:
        engine.dispose()


def test_execute_sql_ddl_create_table_returns_negative_one() -> None:
    """DDL CREATE TABLE 应返回 rowcount=-1（SQLite DDL 不影响行数）."""
    engine = _make_memory_engine()
    try:
        result = execute_sql(engine, "CREATE TABLE foo (id INTEGER PRIMARY KEY)")
        assert result["read_only"] is False
        assert result["rowcount"] == -1
        assert result["columns"] == []
    finally:
        engine.dispose()


def test_execute_sql_read_only_mode_allows_select() -> None:
    """read_only=True 时 SELECT 应正常执行."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = execute_sql(engine, "SELECT 1", read_only=True)
        assert result["read_only"] is True
        assert result["rows"][0].get(1) == 1 or list(result["rows"][0].values()) == [1]
    finally:
        engine.dispose()


def test_execute_sql_read_only_mode_blocks_insert() -> None:
    """read_only=True 时 INSERT 应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="仅允许执行只读"):
            execute_sql(
                engine,
                "INSERT INTO users (name) VALUES ('blocked')",
                read_only=True,
            )
    finally:
        engine.dispose()


def test_execute_sql_read_only_mode_blocks_update() -> None:
    """read_only=True 时 UPDATE 应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="仅允许执行只读"):
            execute_sql(
                engine,
                "UPDATE users SET age = 1 WHERE id = 1",
                read_only=True,
            )
    finally:
        engine.dispose()


def test_execute_sql_read_only_mode_blocks_delete() -> None:
    """read_only=True 时 DELETE 应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="仅允许执行只读"):
            execute_sql(
                engine,
                "DELETE FROM users WHERE id = 1",
                read_only=True,
            )
    finally:
        engine.dispose()


def test_execute_sql_read_only_mode_blocks_ddl() -> None:
    """read_only=True 时 DDL DROP 应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="仅允许执行只读"):
            execute_sql(engine, "DROP TABLE users", read_only=True)
    finally:
        engine.dispose()


def test_execute_sql_empty_raises() -> None:
    """空 SQL 应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError, match="不能为空"):
            execute_sql(engine, "")
    finally:
        engine.dispose()


def test_execute_sql_syntax_error_raises_sqlalchemy_error() -> None:
    """语法错误应抛 SQLAlchemyError."""
    engine = _make_memory_engine()
    try:
        from sqlalchemy.exc import SQLAlchemyError

        with pytest.raises(SQLAlchemyError):
            execute_sql(engine, "SELECT FROM WHERE")
    finally:
        engine.dispose()


def test_execute_sql_write_committed() -> None:
    """DML 写入后应能在新连接读到（事务已提交）."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        execute_sql(
            engine,
            "INSERT INTO users (name, email, age) VALUES ('Gina', 'g@e.com', 28)",
        )
        result = execute_sql(engine, "SELECT COUNT(*) AS cnt FROM users")
        assert result["rows"][0]["cnt"] == 6
    finally:
        engine.dispose()


def test_execute_sql_with_cte_read_only() -> None:
    """WITH 语句应识别为只读."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = execute_sql(
            engine,
            "WITH t AS (SELECT * FROM users WHERE age > 25) SELECT COUNT(*) AS cnt FROM t",
        )
        assert result["read_only"] is True
        assert result["rows"][0]["cnt"] == 3  # Alice/David/Charlie
    finally:
        engine.dispose()


# ============================================================
# P4-3 SQL 查询控制台 - explain_sql
# ============================================================


def test_explain_sql_sqlite_returns_plan() -> None:
    """SQLite EXPLAIN QUERY PLAN 应返回 plan/rows/columns/dialect."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = explain_sql(engine, "SELECT * FROM users WHERE id = 1")
        assert result["dialect"] == "sqlite"
        assert result["analyze"] is False
        assert isinstance(result["plan"], list)
        assert len(result["plan"]) > 0
        assert isinstance(result["rows"], list)
        assert "detail" in result["rows"][0]
        assert "id" in result["columns"]
        assert "parent" in result["columns"]
        assert "notused" in result["columns"]
        assert "detail" in result["columns"]
    finally:
        engine.dispose()


def test_explain_sql_analyze_ignored_on_sqlite() -> None:
    """SQLite 应忽略 analyze 参数."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = explain_sql(engine, "SELECT * FROM users", analyze=True)
        assert result["analyze"] is False
        assert result["dialect"] == "sqlite"
    finally:
        engine.dispose()


def test_explain_sql_empty_raises() -> None:
    """空 SQL 应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError, match="不能为空"):
            explain_sql(engine, "")
    finally:
        engine.dispose()


def test_explain_sql_trailing_semicolon() -> None:
    """末尾分号应被正确处理."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = explain_sql(engine, "SELECT * FROM users;")
        assert len(result["plan"]) > 0
    finally:
        engine.dispose()


def test_explain_sql_unsupported_dialect_raises() -> None:
    """不支持的方言应抛 QueryError（用 FakeEngine 模拟）."""

    class _FakeDialect:
        name = "oracle"

    class _FakeEngine:
        dialect = _FakeDialect()

        def connect(self) -> Any:  # pragma: no cover
            raise AssertionError("不应调用 connect")

    with pytest.raises(QueryError, match="暂不支持 EXPLAIN"):
        explain_sql(cast(Engine, _FakeEngine()), "SELECT 1")


def test_explain_sql_syntax_error_raises_sqlalchemy_error() -> None:
    """SQL 语法错误应抛 SQLAlchemyError."""
    engine = _make_memory_engine()
    try:
        from sqlalchemy.exc import SQLAlchemyError

        with pytest.raises(SQLAlchemyError):
            explain_sql(engine, "SELECT FROM WHERE")
    finally:
        engine.dispose()


# ============================================================
# P4-4 导入导出 - _format_csv_value / _format_sql_value / _format_excel_value
# ============================================================


def test_format_csv_value_none() -> None:
    """None 应转为空字符串."""
    assert _format_csv_value(None) == ""


def test_format_csv_value_bool() -> None:
    """bool 应转为 '1'/'0'."""
    assert _format_csv_value(True) == "1"
    assert _format_csv_value(False) == "0"


def test_format_csv_value_numbers() -> None:
    """int/float 应转为字符串."""
    assert _format_csv_value(42) == "42"
    assert _format_csv_value(3.14) == "3.14"


def test_format_csv_value_datetime() -> None:
    """datetime/date/time 应转为 ISO 格式字符串."""
    dt = datetime(2026, 7, 31, 12, 30, 45)
    assert _format_csv_value(dt) == "2026-07-31T12:30:45"
    d = date(2026, 7, 31)
    assert _format_csv_value(d) == "2026-07-31"


def test_format_csv_value_bytes() -> None:
    """bytes 应 UTF-8 解码（失败用 replace）."""
    assert _format_csv_value(b"hello") == "hello"
    assert _format_csv_value(b"\xff\xfe") == "\ufffd\ufffd"


def test_format_csv_value_dict_list() -> None:
    """dict/list 应转为 JSON 字符串."""
    assert _format_csv_value({"a": 1}) == '{"a": 1}'
    assert _format_csv_value([1, 2, 3]) == "[1, 2, 3]"


def test_format_csv_value_str() -> None:
    """普通字符串应原样返回."""
    assert _format_csv_value("hello") == "hello"


def test_format_sql_value_none() -> None:
    """None 应转为 NULL."""
    assert _format_sql_value(None, "sqlite") == "NULL"


def test_format_sql_value_bool() -> None:
    """bool 应转为 1/0."""
    assert _format_sql_value(True, "sqlite") == "1"
    assert _format_sql_value(False, "sqlite") == "0"


def test_format_sql_value_numbers() -> None:
    """int/float 应转为数字字面量."""
    assert _format_sql_value(42, "sqlite") == "42"
    assert _format_sql_value(3.14, "sqlite") == "3.14"


def test_format_sql_value_datetime() -> None:
    """datetime 应转为 'ISO' 字符串."""
    dt = datetime(2026, 7, 31, 12, 30, 45)
    assert _format_sql_value(dt, "sqlite") == "'2026-07-31T12:30:45'"


def test_format_sql_value_bytes_sqlite() -> None:
    """SQLite bytes 应转为 X'hex'."""
    assert _format_sql_value(b"\xab\xcd", "sqlite") == "X'abcd'"


def test_format_sql_value_bytes_other_dialect() -> None:
    """非 SQLite bytes 应转为 'utf-8 解码' 字符串."""
    assert _format_sql_value(b"hello", "postgresql") == "'hello'"


def test_format_sql_value_str_escape() -> None:
    """字符串单引号应翻倍转义."""
    assert _format_sql_value("hello", "sqlite") == "'hello'"
    assert _format_sql_value("it's", "sqlite") == "'it''s'"


def test_format_excel_value_basic() -> None:
    """Excel 值应保留原始类型（str/int/float/bool/datetime/None）."""
    assert _format_excel_value(None) is None
    assert _format_excel_value(True) is True
    assert _format_excel_value(42) == 42
    assert _format_excel_value(3.14) == 3.14
    assert _format_excel_value("hello") == "hello"
    dt = datetime(2026, 7, 31)
    assert _format_excel_value(dt) == dt


def test_format_excel_value_bytes_dict() -> None:
    """bytes/dict 应转为可序列化值."""
    assert _format_excel_value(b"hello") == "hello"
    assert _format_excel_value({"a": 1}) == '{"a": 1}'


# ============================================================
# P4-4 导入导出 - iter_table_rows
# ============================================================


def test_iter_table_rows_streams_all_rows() -> None:
    """应流式生成所有行."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows = list(iter_table_rows(engine, "users", schema=None))
        assert len(rows) == 5
        assert rows[0]["name"] == "Alice"
    finally:
        engine.dispose()


def test_iter_table_rows_batch_size() -> None:
    """小 batch_size 也应正确分批."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows = list(iter_table_rows(engine, "users", schema=None, batch_size=2))
        assert len(rows) == 5
    finally:
        engine.dispose()


def test_iter_table_rows_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            list(iter_table_rows(engine, "nonexistent", schema=None))
    finally:
        engine.dispose()


# ============================================================
# P4-4 导入导出 - rows_to_csv
# ============================================================


def test_rows_to_csv_header_and_rows() -> None:
    """应首行输出 BOM+表头，后续每行一条记录."""
    rows = iter([{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}])
    chunks = list(rows_to_csv(rows, ["id", "name"]))
    # 首块含 BOM + 表头
    assert chunks[0].startswith("\ufeff")
    assert "id,name" in chunks[0]
    # 后续为数据行
    assert any("1,Alice" in c for c in chunks)
    assert any("2,Bob" in c for c in chunks)


def test_rows_to_csv_none_to_empty() -> None:
    """None 值应转为空字符串."""
    rows = iter([{"id": 1, "name": None}])
    chunks = list(rows_to_csv(rows, ["id", "name"]))
    assert any("1," in c and "Alice" not in c for c in chunks[1:])


# ============================================================
# P4-4 导入导出 - rows_to_sql
# ============================================================


def test_rows_to_sql_insert_statements() -> None:
    """应生成 INSERT 语句."""
    rows = iter([{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}])
    chunks = list(rows_to_sql(rows, ["id", "name"], "users", None, "sqlite"))
    assert all(c.startswith('INSERT INTO "users" ("id", "name") VALUES ') for c in chunks)
    assert chunks[0].endswith(");\n")
    assert "1, 'Alice'" in chunks[0]
    assert "2, 'Bob'" in chunks[1]


def test_rows_to_sql_none_to_null() -> None:
    """None 应转为 NULL."""
    rows = iter([{"id": 1, "name": None}])
    chunks = list(rows_to_sql(rows, ["id", "name"], "users", None, "sqlite"))
    assert "1, NULL" in chunks[0]


def test_rows_to_sql_escape_quote() -> None:
    """单引号应翻倍转义."""
    rows = iter([{"id": 1, "name": "it's"}])
    chunks = list(rows_to_sql(rows, ["id", "name"], "users", None, "sqlite"))
    assert "'it''s'" in chunks[0]


# ============================================================
# P4-4 导入导出 - export_excel
# ============================================================


def test_export_excel_returns_xlsx_bytes() -> None:
    """应返回可被 openpyxl 解析的 xlsx bytes."""
    from openpyxl import load_workbook

    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        data = export_excel(engine, "users", schema=None)
        assert isinstance(data, bytes)
        assert data[:2] == b"PK"  # xlsx 是 zip 格式，签名以 PK 开头
        # 用 openpyxl 读取验证
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws is not None
        rows = list(ws.iter_rows(values_only=True))
        # 表头 + 5 行
        assert rows[0] == ("id", "name", "email", "age")
        assert len(rows) == 6
        assert rows[1][1] == "Alice"
        wb.close()
    finally:
        engine.dispose()


def test_export_excel_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            export_excel(engine, "nonexistent", schema=None)
    finally:
        engine.dispose()


# ============================================================
# P4-4 导入导出 - parse_csv_upload
# ============================================================


def test_parse_csv_upload_basic() -> None:
    """应返回表头与行 dict 迭代器."""
    content = b"id,name,age\n1,Alice,30\n2,Bob,25\n"
    file_obj = io.BytesIO(content)
    headers, rows_iter = parse_csv_upload(file_obj)
    assert headers == ["id", "name", "age"]
    rows = list(rows_iter)
    assert rows == [
        {"id": "1", "name": "Alice", "age": "30"},
        {"id": "2", "name": "Bob", "age": "25"},
    ]


def test_parse_csv_upload_with_bom() -> None:
    """应正确处理 UTF-8 BOM."""
    content = b"\xef\xbb\xbfid,name\n1,Alice\n"
    file_obj = io.BytesIO(content)
    headers, _ = parse_csv_upload(file_obj)
    assert headers == ["id", "name"]


def test_parse_csv_upload_empty_raises() -> None:
    """空文件应抛 QueryError."""
    file_obj = io.BytesIO(b"")
    with pytest.raises(QueryError, match="CSV 文件为空"):
        parse_csv_upload(file_obj)


def test_parse_csv_upload_chinese() -> None:
    """应正确解析中文."""
    content = "id,name\n1,张三\n2,李四\n".encode()
    file_obj = io.BytesIO(content)
    _headers, rows_iter = parse_csv_upload(file_obj)
    rows = list(rows_iter)
    assert rows[0]["name"] == "张三"
    assert rows[1]["name"] == "李四"


# ============================================================
# P4-4 导入导出 - parse_excel_upload
# ============================================================


def _make_xlsx_bytes(headers: list[str], data_rows: list[list[Any]]) -> bytes:
    """构造 xlsx bytes 用于测试."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_excel_upload_basic() -> None:
    """应返回表头与行 dict 迭代器."""
    data = _make_xlsx_bytes(["id", "name"], [[1, "Alice"], [2, "Bob"]])
    file_obj = io.BytesIO(data)
    headers, rows_iter = parse_excel_upload(file_obj)
    assert headers == ["id", "name"]
    rows = list(rows_iter)
    assert rows[0] == {"id": 1, "name": "Alice"}
    assert rows[1] == {"id": 2, "name": "Bob"}


def test_parse_excel_upload_empty_raises() -> None:
    """空工作表应抛 QueryError."""
    # 仅创建空工作表（无任何行）
    from openpyxl import Workbook

    wb = Workbook()
    # 删除默认 sheet 后新建一个真正空的 sheet
    default_ws = wb.active
    assert default_ws is not None
    wb.remove(default_ws)
    wb.create_sheet()
    buf = io.BytesIO()
    wb.save(buf)
    file_obj = io.BytesIO(buf.getvalue())
    with pytest.raises(QueryError, match="Excel 文件为空"):
        parse_excel_upload(file_obj)


# ============================================================
# P4-4 导入导出 - import_rows
# ============================================================


def test_import_rows_success() -> None:
    """应批量插入所有行."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows = iter(
            [
                {"id": 10, "name": "X", "email": "x@e.com", "age": 20},
                {"id": 11, "name": "Y", "email": "y@e.com", "age": 21},
            ]
        )
        result = import_rows(engine, "users", None, ["id", "name", "email", "age"], rows)
        assert result["success_count"] == 2
        assert result["failed_count"] == 0
        assert result["errors"] == []
        # 验证数据已写入
        _rows_db, total = query_table_rows(engine, "users", schema=None)
        assert total == 7
    finally:
        engine.dispose()


def test_import_rows_empty_columns_raises() -> None:
    """空列名列表应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="列名列表不能为空"):
            import_rows(engine, "users", None, [], iter([]))
    finally:
        engine.dispose()


def test_import_rows_invalid_column_raises() -> None:
    """非法列名应抛 QueryError."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        with pytest.raises(QueryError, match="非法列名"):
            import_rows(engine, "users", None, ["nonexistent"], iter([{"nonexistent": 1}]))
    finally:
        engine.dispose()


def test_import_rows_constraint_violation_rolls_back() -> None:
    """约束冲突应抛 SQLAlchemyError 且事务回滚（无部分插入）."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        # 第二行 name 为 NOT NULL，传 None 触发约束冲突
        rows = iter(
            [
                {"id": 10, "name": "X", "email": "x@e.com", "age": 20},
                {"id": 11, "name": None, "email": "y@e.com", "age": 21},
            ]
        )
        from sqlalchemy.exc import SQLAlchemyError

        with pytest.raises(SQLAlchemyError):
            import_rows(
                engine,
                "users",
                None,
                ["id", "name", "email", "age"],
                rows,
            )
        # 事务回滚：第一行也不应存在
        rows_db, total = query_table_rows(engine, "users", schema=None)
        assert total == 5
        assert all(r["id"] != 10 for r in rows_db)
    finally:
        engine.dispose()


def test_import_rows_unknown_table_raises() -> None:
    """不存在的表应抛 QueryError（get_column_names 反射失败）."""
    engine = _make_memory_engine()
    try:
        with pytest.raises(QueryError):
            import_rows(engine, "nonexistent", None, ["id"], iter([{"id": 1}]))
    finally:
        engine.dispose()


def test_import_rows_missing_column_fills_null() -> None:
    """缺失列应填 NULL（数据库默认值生效或 NOT NULL 冲突抛错）."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        # 仅提供 id 与 name，email/age 缺失（age 有默认值，email 可空）
        rows = iter([{"id": 10, "name": "X"}])
        result = import_rows(engine, "users", None, ["id", "name"], rows)
        assert result["success_count"] == 1
        # 验证 email 为 None，age 为默认值 0
        row = get_row(engine, "users", schema=None, pk={"id": 10})
        assert row is not None
        assert row["email"] is None
        assert row["age"] == 0
    finally:
        engine.dispose()


# ============================================================
# P4-4 补充：_format_excel_value fallback 分支（覆盖 query.py:835）
# ============================================================


def test_format_excel_value_fallback_to_str() -> None:
    """非预期类型（如 set）应 fallback 到 str(val)."""
    result = _format_excel_value({1, 2, 3})
    assert isinstance(result, str)
    assert "1" in result


# ============================================================
# P4-4 补充：parse_csv_upload 非 bytes 分支（覆盖 query.py:997-999）
# ============================================================


class _StringFileObj:
    """模拟返回字符串（非 bytes）的文件对象."""

    def __init__(self, content: str) -> None:
        self._content = content

    def read(self) -> str:
        return self._content


class _StringFileObjWithBom:
    """模拟返回带 BOM 字符串的文件对象."""

    def __init__(self, content: str) -> None:
        self._content = content

    def read(self) -> str:
        return self._content


def test_parse_csv_upload_string_input() -> None:
    """字符串输入（非 bytes）应正常解析."""
    content = "id,name,age\n1,Alice,30\n2,Bob,25\n"
    file_obj = _StringFileObj(content)
    headers, rows_iter = parse_csv_upload(file_obj)
    assert headers == ["id", "name", "age"]
    rows = list(rows_iter)
    assert len(rows) == 2
    assert rows[0]["name"] == "Alice"


def test_parse_csv_upload_string_with_bom() -> None:
    """字符串输入含 BOM 应正确去除."""
    content = "\ufeffid,name\n1,Alice\n"
    file_obj = _StringFileObjWithBom(content)
    headers, _ = parse_csv_upload(file_obj)
    assert headers == ["id", "name"]


# ============================================================
# P4-4 补充：import_rows 批量插入分支（覆盖 query.py:1110-1117）
# ============================================================


def test_import_rows_batch_insert_when_batch_full() -> None:
    """batch_size 触发批量提交分支应正常工作."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows = iter(
            [
                {"id": 10, "name": "A", "email": "a@e.com", "age": 20},
                {"id": 11, "name": "B", "email": "b@e.com", "age": 21},
                {"id": 12, "name": "C", "email": "c@e.com", "age": 22},
            ]
        )
        # batch_size=2 触发 "if len(batch) >= batch_size" 分支
        result = import_rows(
            engine,
            "users",
            None,
            ["id", "name", "email", "age"],
            rows,
            batch_size=2,
        )
        assert result["success_count"] == 3
        _rows_db, total = query_table_rows(engine, "users", schema=None)
        assert total == 8  # 原 5 + 3
    finally:
        engine.dispose()


def test_import_rows_empty_rows_iter() -> None:
    """空行迭代器应返回 success_count=0."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        result = import_rows(engine, "users", None, ["id", "name"], iter([]))
        assert result["success_count"] == 0
        assert result["failed_count"] == 0
        _rows_db, total = query_table_rows(engine, "users", schema=None)
        assert total == 5  # 不变
    finally:
        engine.dispose()


# ============================================================
# P4-6 大数据量流式测试（标记为 slow，需显式运行）
# ============================================================


@pytest.mark.slow
def test_iter_table_rows_1000_plus_rows() -> None:
    """流式读取 1000+ 行应全部返回且内存效率可接受."""
    engine = _make_memory_engine()
    try:
        with engine.begin() as conn:
            conn.execute(
                text(
                    "CREATE TABLE big ("
                    "id INTEGER PRIMARY KEY AUTOINCREMENT, "
                    "name VARCHAR(50) NOT NULL, "
                    "value REAL DEFAULT 0"
                    ")"
                )
            )
            rows = [{"name": f"item_{i}", "value": i * 1.5} for i in range(1200)]
            conn.execute(
                text("INSERT INTO big (name, value) VALUES (:name, :value)"),
                rows,
            )
        result = list(iter_table_rows(engine, "big", schema=None, batch_size=100))
        assert len(result) == 1200
        assert result[0]["name"] == "item_0"
        assert result[-1]["name"] == "item_1199"
    finally:
        engine.dispose()


@pytest.mark.slow
def test_import_rows_large_batch_performance() -> None:
    """批量导入 500 行应在合理时间内完成."""
    engine = _make_memory_engine()
    try:
        _setup_tables(engine)
        rows = [
            {"id": 100 + i, "name": f"user_{i}", "email": f"user_{i}@example.com", "age": 20 + i % 50}
            for i in range(500)
        ]
        result = import_rows(
            engine,
            "users",
            None,
            ["id", "name", "email", "age"],
            iter(rows),
            batch_size=100,
        )
        assert result["success_count"] == 500
        _rows_db, total = query_table_rows(engine, "users", schema=None)
        assert total == 505  # 原 5 + 500
    finally:
        engine.dispose()


@pytest.mark.slow
def test_export_excel_1000_rows() -> None:
    """导出 1000 行应为有效的 xlsx 文件."""
    from openpyxl import load_workbook

    engine = _make_memory_engine()
    try:
        with engine.begin() as conn:
            conn.execute(text("CREATE TABLE xport (id INTEGER PRIMARY KEY AUTOINCREMENT, name VARCHAR(50))"))
            rows = [{"name": f"row_{i}"} for i in range(1000)]
            conn.execute(
                text("INSERT INTO xport (name) VALUES (:name)"),
                rows,
            )
        data = export_excel(engine, "xport", schema=None)
        assert isinstance(data, bytes)
        assert data[:2] == b"PK"
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws is not None
        all_rows = list(ws.iter_rows(values_only=True))
        assert len(all_rows) == 1001  # 表头 + 1000 行
        wb.close()
    finally:
        engine.dispose()


@pytest.mark.slow
def test_rows_to_csv_large_stream() -> None:
    """大数据量 CSV 流式生成器应逐块产出."""
    rows = iter([{"id": i, "name": f"user_{i}"} for i in range(500)])
    chunks = list(rows_to_csv(rows, ["id", "name"]))
    assert len(chunks) == 501  # 1 表头块 + 500 数据块
    # 首块含 BOM + 表头
    assert chunks[0].startswith("\ufeff")
    assert "id,name" in chunks[0]
    # 末块含最后一行
    assert "499,user_499" in chunks[-1]


@pytest.mark.slow
def test_rows_to_sql_large_stream() -> None:
    """大数据量 SQL 流式生成器应逐行产出 INSERT."""
    rows = iter([{"id": i, "name": f"user_{i}"} for i in range(500)])
    chunks = list(rows_to_sql(rows, ["id", "name"], "users", None, "sqlite"))
    assert len(chunks) == 500
    # 每行应以分号结尾
    assert all(c.endswith(";\n") for c in chunks)
    # 首行和末行
    assert "0," in chunks[0]
    assert "user_0" in chunks[0]
    assert "499," in chunks[-1]
    assert "user_499" in chunks[-1]
